# -*- coding: utf-8 -*-
"""Add fifteen wineries to the marketing workbook and top up three existing ones.

Rows are written in the sheets' own styles, new winery blocks are appended as
whole groups, top-up rows are spliced next to the winery they belong to, and
every COUNTIF/COUNTIFS range that pointed at an old last row is widened.
"""
import shutil
from copy import copy

import openpyxl
from openpyxl.utils import get_column_letter

from research import WINERIES, TOPUP_POSITIONS, TOPUP_CAMPAIGNS, CHECK

SRC = "Romanian_Wineries_Marketing_finalfinal.xlsx"
OUT = "Romanian_Wineries_Marketing_finalfinal.xlsx"

wb = openpyxl.load_workbook(SRC)


def styles(ws, row, ncols, c0=1):
    return [copy(ws.cell(row=row, column=c)._style) for c in range(c0, c0 + ncols)]


def apply(ws, row, st, ncols, c0=1):
    for i, c in enumerate(range(c0, c0 + ncols)):
        ws.cell(row=row, column=c)._style = copy(st[i])


def splice(ws, ncols, appends, topups):
    """appends: list of row-lists added as new groups at the end.
       topups: list of row-lists spliced after the last row of the same winery."""
    old_last = ws.max_row
    data = [[ws.cell(row=r, column=c).value for c in range(1, ncols + 1)]
            for r in range(2, old_last + 1)]
    st_first = styles(ws, 2, ncols)
    st_mid = styles(ws, (2 + old_last) // 2, ncols)
    st_last = styles(ws, old_last, ncols)

    for row in topups:
        idx = max(i for i, r in enumerate(data) if r[0] == row[0])
        data.insert(idx + 1, row)
    data.extend(appends)

    new_last = 1 + len(data)
    for i, row in enumerate(data):
        r = 2 + i
        for c in range(1, ncols + 1):
            ws.cell(row=r, column=c).value = row[c - 1]
        apply(ws, r, st_mid, ncols)
    apply(ws, 2, st_first, ncols)
    apply(ws, new_last, st_last, ncols)
    return old_last, new_last


def widen(ws, coords, pairs):
    for coord in coords:
        v = ws[coord].value
        if isinstance(v, str) and v.startswith("="):
            for a, b in pairs:
                v = v.replace(a, b)
            ws[coord] = v


# ------------------------------------------------------------ Open Positions
op_new_rows = []
for w in WINERIES:
    for p in w["positions"]:
        op_new_rows.append([w["name"]] + p)
op_old, op_last = splice(wb["Open Positions"], 8, op_new_rows, TOPUP_POSITIONS)

# --------------------------------------------------------- Campaign Evidence
ce_new_rows = []
for w in WINERIES:
    for c in w["campaigns"]:
        ce_new_rows.append([w["name"]] + c)
ce_old, ce_last = splice(wb["Campaign Evidence"], 7, ce_new_rows, TOPUP_CAMPAIGNS)

# --------------------------------------------------------------- CEO Research
ceo_rows = [[w["name"]] + w["ceo"] for w in WINERIES]
ceo_old, ceo_last = splice(wb["CEO Research"], 8, ceo_rows, [])

# ----------------------------------------------------------- Commercial People
cp = wb["Commercial People"]
cp_rows = []
for w in WINERIES:
    for p in w["people"]:
        cp_rows.append([w["name"]] + p)
cp_old, cp_last = splice(cp, 7, cp_rows, [])
cp_pairs = [("$C$2:$C$88", f"$C$2:$C${cp_last}"), ("$E$2:$E$88", f"$E$2:$E${cp_last}")]
widen(cp, ["J2", "J3"] + [f"J{r}" for r in range(7, 14)] + [f"M{r}" for r in range(7, 18)], cp_pairs)

# ------------------------------------------------------- Commercial Type Split
ts = wb["Commercial Type Split"]
for rng in list(ts.merged_cells.ranges):
    if str(rng) == "A20:V20":
        ts.unmerge_cells("A20:V20")
note_style = copy(ts.cell(row=20, column=1)._style)
ts.cell(row=20, column=1).value = None

TOTAL_ROW_OLD = 18
st_data = styles(ts, 3, 22)
st_dlast = styles(ts, 17, 22)
st_total = styles(ts, TOTAL_ROW_OLD, 22)
total_vals = [ts.cell(row=TOTAL_ROW_OLD, column=c).value for c in range(1, 23)]
for c in range(1, 23):
    ts.cell(row=TOTAL_ROW_OLD, column=c).value = None
    ts.cell(row=TOTAL_ROW_OLD, column=c)._style = copy(st_data[c - 1])

first_new = TOTAL_ROW_OLD                      # 18
for i, w in enumerate(WINERIES):
    r = first_new + i
    ts.cell(row=r, column=1).value = w["name"]
    for c in range(2, 9):                      # B..H marketing types
        col = get_column_letter(c)
        ts.cell(row=r, column=c).value = (
            f"=COUNTIFS('Commercial People'!$A$2:$A${cp_last},\"{w['name']}\","
            f"'Commercial People'!$C$2:$C${cp_last},\"Marketing\","
            f"'Commercial People'!$E$2:$E${cp_last},{col}$2)")
    ts.cell(row=r, column=9).value = f"=SUM(B{r}:H{r})"
    for c in range(10, 21):                    # J..T sales types
        col = get_column_letter(c)
        ts.cell(row=r, column=c).value = (
            f"=COUNTIFS('Commercial People'!$A$2:$A${cp_last},\"{w['name']}\","
            f"'Commercial People'!$C$2:$C${cp_last},\"Sales\","
            f"'Commercial People'!$E$2:$E${cp_last},{col}$2)")
    ts.cell(row=r, column=21).value = f"=SUM(J{r}:T{r})"
    ts.cell(row=r, column=22).value = f"=I{r}+U{r}"
    apply(ts, r, st_data, 22)
    ts.row_dimensions[r].height = 22.5

ts_data_last = first_new + len(WINERIES) - 1
apply(ts, ts_data_last, st_dlast, 22)
ts_total = ts_data_last + 1
for c in range(1, 23):
    v = total_vals[c - 1]
    if isinstance(v, str) and v.startswith("=SUM("):
        col = get_column_letter(c)
        v = f"=SUM({col}3:{col}{ts_data_last})"
    ts.cell(row=ts_total, column=c).value = v
apply(ts, ts_total, st_total, 22)
ts.row_dimensions[ts_total].height = 22.5

ts_pairs = [(f"$A$2:$A$88", f"$A$2:$A${cp_last}"),
            (f"$C$2:$C$88", f"$C$2:$C${cp_last}"),
            (f"$E$2:$E$88", f"$E$2:$E${cp_last}")]
widen(ts, [f"{get_column_letter(c)}{r}" for r in range(3, TOTAL_ROW_OLD)
           for c in list(range(2, 9)) + list(range(10, 21))], ts_pairs)

note_row = ts_total + 2
ts.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=22)
ts.cell(row=note_row, column=1).value = (
    "Counts are formula-driven from the Commercial People sheet. Wineries added on "
    f"{CHECK} are researched from public sources only — company team pages, trade press and "
    "job boards — because the LinkedIn capture behind the original rows could not be "
    "reproduced. Where no marketing or sales person is named publicly, the row reads zero: "
    "that means no public evidence was found, not that the function is unstaffed.")
ts.cell(row=note_row, column=1)._style = copy(note_style)
ts.row_dimensions[note_row].height = 30.0

# ----------------------------------------------------------------- Takeaways
tk = wb["Takeaways_Insights"]
tk_rows = []
for w in WINERIES:
    src = w["bestjobs"][0]
    freshness = (f"Public sources checked {CHECK}; no LinkedIn capture"
                 if src else f"Public sources checked {CHECK}")
    t = w["takeaway"]
    tk_rows.append([w["name"], None, freshness, None, t[0], t[1], t[2], t[3], t[4]])
tk_old, tk_last = splice(tk, 9, tk_rows, [])

for r in range(2, tk_last + 1):
    cur = tk.cell(row=r, column=4).value
    if not (isinstance(cur, str) and cur.startswith("=")):
        tk.cell(row=r, column=4).value = (
            f"=COUNTIFS('Open Positions'!$A$2:$A${op_last},A{r},"
            f"'Open Positions'!$C$2:$C${op_last},\"Vacancy\")"
            f"+COUNTIFS('Open Positions'!$A$2:$A${op_last},A{r},"
            f"'Open Positions'!$C$2:$C${op_last},\"Paid internship / early career\")")
widen(tk, [f"D{r}" for r in range(2, tk_last + 1)],
      [("$A$2:$A$29", f"$A$2:$A${op_last}"), ("$C$2:$C$29", f"$C$2:$C${op_last}")])

# --------------------------------------------------------- Winery Departments
wd = wb["Winery Departments"]
wd_last_summary = max(r for r in range(2, wd.max_row + 1) if wd.cell(row=r, column=5).value)
st_sum = styles(wd, 3, 5, c0=5)
st_sum_last = styles(wd, wd_last_summary, 5, c0=5)
apply(wd, wd_last_summary, st_sum, 5, c0=5)
for i, w in enumerate(WINERIES):
    r = wd_last_summary + 1 + i
    wd.cell(row=r, column=5).value = w["name"]
    wd.cell(row=r, column=6).value = None
    wd.cell(row=r, column=7).value = (
        f"=SUMIF('Winery Departments'!$A:$A,E{r},'Winery Departments'!$C:$C)")
    wd.cell(row=r, column=8).value = f'=IFERROR(G{r}/F{r},"")'
    wd.cell(row=r, column=9).value = w["dept_note"]
    apply(wd, r, st_sum, 5, c0=5)
apply(wd, wd_last_summary + len(WINERIES), st_sum_last, 5, c0=5)

wb.save(OUT)
print(f"Open Positions      {op_old} -> {op_last}")
print(f"Campaign Evidence   {ce_old} -> {ce_last}")
print(f"CEO Research        {ceo_old} -> {ceo_last}")
print(f"Commercial People   {cp_old} -> {cp_last}")
print(f"Type Split          data 3-{ts_data_last}, total {ts_total}, note {note_row}")
print(f"Takeaways           {tk_old} -> {tk_last}")
print(f"Winery Departments  summary rows 2-{wd_last_summary + len(WINERIES)}")
