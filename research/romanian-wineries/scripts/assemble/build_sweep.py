# -*- coding: utf-8 -*-
"""Add the 2021-2026 sweep to Campaign Evidence and build the Campaign Timeline sheet."""
import json
from copy import copy

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from sweep_rows import SWEEP

FILE = "Romanian_Wineries_Marketing_finalfinal.xlsx"
wb = openpyxl.load_workbook(FILE)

# ------------------------------------------------- 1. Campaign Evidence -----
ce = wb["Campaign Evidence"]
NC = 7
old_last = ce.max_row
data = [[ce.cell(row=r, column=c).value for c in range(1, NC + 1)] for r in range(2, old_last + 1)]
st_first = [copy(ce.cell(row=2, column=c)._style) for c in range(1, NC + 1)]
st_mid = [copy(ce.cell(row=60, column=c)._style) for c in range(1, NC + 1)]
st_last = [copy(ce.cell(row=old_last, column=c)._style) for c in range(1, NC + 1)]

known = {row[0] for row in data}
missing = sorted({r[0] for r in SWEEP} - known)
if missing:
    raise SystemExit(f"winery name not found in Campaign Evidence: {missing}")

for row in SWEEP:
    idx = max(i for i, d in enumerate(data) if d[0] == row[0])
    data.insert(idx + 1, list(row))

new_last = 1 + len(data)
for i, row in enumerate(data):
    r = 2 + i
    for c in range(1, NC + 1):
        ce.cell(row=r, column=c).value = row[c - 1]
    for c in range(1, NC + 1):
        ce.cell(row=r, column=c)._style = copy(st_mid[c - 1])
for c in range(1, NC + 1):
    ce.cell(row=2, column=c)._style = copy(st_first[c - 1])
    ce.cell(row=new_last, column=c)._style = copy(st_last[c - 1])
print(f"Campaign Evidence {old_last} -> {new_last}  (+{len(SWEEP)} sweep rows)")

# -------------------------------------------------- 2. Campaign Timeline ----
tl = json.load(open("timeline.json"))
if "Campaign Timeline" in wb.sheetnames:
    del wb["Campaign Timeline"]
ws = wb.create_sheet("Campaign Timeline", wb.sheetnames.index("Campaign Evidence") + 1)

NAVY = "FF17365D"
hdr_font = Font(name="Arial", size=9, bold=True, color="FFFFFFFF")
hdr_fill = PatternFill("solid", fgColor=NAVY)
thin = Side(style="thin", color="FFBFBFBF")
med = Side(style="medium", color="FF17365D")
body_font = Font(name="Arial", size=9, color="FF000000")
bold_font = Font(name="Arial", size=9, bold=True, color="FF000000")
url_font = Font(name="Arial", size=8, color="FFC00000")
white = PatternFill("solid", fgColor="FFFFFFFF")

HEAD = ["Winery", "Year", "Month", "Date", "Campaign / item", "Type",
        "Where it was found", "Source URL"]
WIDTH = [24, 7, 8, 12, 68, 24, 34, 60]
for c, (h, w) in enumerate(zip(HEAD, WIDTH), start=1):
    cell = ws.cell(row=1, column=c, value=h)
    cell.font = hdr_font
    cell.fill = hdr_fill
    cell.alignment = Alignment(wrap_text=True, vertical="center")
    cell.border = Border(bottom=med)
    ws.column_dimensions[get_column_letter(c)].width = w
ws.row_dimensions[1].height = 24
ws.freeze_panes = "A2"

for i, r in enumerate(tl):
    row = 2 + i
    vals = [r["winery"], int(r["year"]), (r["month"] or ""), (r["date"] or ""),
            r["title"], r["type"], r["source"], r["url"]]
    for c, v in enumerate(vals, start=1):
        cell = ws.cell(row=row, column=c, value=v)
        cell.font = bold_font if c == 1 else (url_font if c == 8 else body_font)
        cell.fill = white
        cell.alignment = Alignment(wrap_text=(c in (5, 6, 7, 8)), vertical="top")
        cell.border = Border(bottom=thin)
        if c == 2:
            cell.number_format = '0'
            cell.alignment = Alignment(horizontal="center", vertical="top")
        if c == 3:
            cell.alignment = Alignment(horizontal="center", vertical="top")
last = 1 + len(tl)
ws.auto_filter.ref = f"A1:H{last}"

note = last + 2
ws.merge_cells(start_row=note, start_column=1, end_row=note, end_column=8)
nc = ws.cell(row=note, column=1)
nc.value = (
    "How this sheet was built (sweep run 2026-08-21). Every winery's own site was crawled for its news, blog or events "
    "section, using the WordPress and Shopify feeds where they exist and the served HTML where they do not; the Romanian "
    "marketing trade press (IQads) was swept in parallel by brand. Only items published from 1 January 2021 onward are "
    "listed, each with the page it came from. Rows whose Date is blank carry a Year taken from the publisher's article "
    "sequence, calibrated against four articles with known dates — treat those years as approximate. "
    "Coverage is uneven by design, not by sampling: seventeen wineries publish dated news that can be read from outside, "
    "and the remaining thirteen publish none, so their campaigns appear in Campaign Evidence from press and agency "
    "sources but cannot be dated month by month here. A low row count means the winery publishes little, not that it "
    "does little.")
nc.font = Font(name="Arial", size=8, color="FF666666")
nc.alignment = Alignment(wrap_text=True, vertical="top")
ws.row_dimensions[note].height = 72

# per-winery coverage summary, to the right of the data
sc = 10
ws.cell(row=1, column=sc, value="Coverage by winery").font = hdr_font
ws.cell(row=1, column=sc).fill = hdr_fill
ws.cell(row=1, column=sc + 1, value="Rows").font = hdr_font
ws.cell(row=1, column=sc + 1).fill = hdr_fill
for c in (sc, sc + 1):
    ws.cell(row=1, column=c).border = Border(bottom=med)
    ws.cell(row=1, column=c).alignment = Alignment(wrap_text=True, vertical="center")
ws.column_dimensions[get_column_letter(sc)].width = 26
ws.column_dimensions[get_column_letter(sc + 1)].width = 8

wineries = sorted({ce.cell(row=r, column=1).value for r in range(2, new_last + 1)})
for i, w in enumerate(wineries):
    r = 2 + i
    ws.cell(row=r, column=sc, value=w).font = bold_font
    f = ws.cell(row=r, column=sc + 1,
                value=f'=COUNTIF($A$2:$A${last},{get_column_letter(sc)}{r})')
    f.font = body_font
    f.number_format = '#,##0;[Red](#,##0);-'
    f.alignment = Alignment(horizontal="right")
    for c in (sc, sc + 1):
        ws.cell(row=r, column=c).border = Border(bottom=thin)
        ws.cell(row=r, column=c).fill = white
tot = 2 + len(wineries)
ws.cell(row=tot, column=sc, value="Total").font = bold_font
t = ws.cell(row=tot, column=sc + 1,
            value=f'=SUM({get_column_letter(sc+1)}2:{get_column_letter(sc+1)}{tot-1})')
t.font = bold_font
t.number_format = '#,##0;[Red](#,##0);-'
t.alignment = Alignment(horizontal="right")
for c in (sc, sc + 1):
    ws.cell(row=tot, column=c).border = Border(top=Side(style="medium"), bottom=Side(style="double"))

wb.save(FILE)
print(f"Campaign Timeline: {len(tl)} rows, note at {note}, coverage block rows 2-{tot}")
