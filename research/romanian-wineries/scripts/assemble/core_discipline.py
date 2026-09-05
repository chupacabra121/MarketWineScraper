# -*- coding: utf-8 -*-
"""Add a standardized "Core discipline" column to CEO Research.

One value per leader: the professional discipline they trained in or built
their career on before wine.  Where a leader holds more than one, education is
the tie-breaker; where no leader is identified the row reads "Not established".
"""
from copy import copy

import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

F = "Romanian_Wineries_Marketing_finalfinal.xlsx"
COL = 7                      # inserted after "Career domains"
HEADER = "Core discipline"

VOCAB = ["Wine Engineer", "Agronomy / Horticulture", "Engineering", "Finance / Economics",
         "Marketing / Sales", "Law", "Medicine", "Academia", "Entrepreneur / Owner",
         "General Management", "Not established"]

# winery -> core discipline
CORE = {
 "Jidvei":                    "Engineering",
 "Beciul Domnesc":            "Finance / Economics",
 "Purcari Wineries":          "Wine Engineer",
 "Cramele Recas":             "Agronomy / Horticulture",
 "Via Viticola":              "Academia",
 "Budureasca":                "Finance / Economics",
 "Zarea":                     "Engineering",
 "Cramele Cotnari":           "Marketing / Sales",
 "Casa de Vinuri Cotnari SA": "Law",
 "Murfatlar":                 "Finance / Economics",
 "Domeniile Averesti":        "Finance / Economics",
 "Tohani Romania":            "Entrepreneur / Owner",
 "Domeniile Ostrov":          "Entrepreneur / Owner",
 "Crama Ceptura":             "Medicine",
 "Davino Winery":             "Entrepreneur / Owner",
 "Crama Gîrboiu":             "Entrepreneur / Owner",
 "Domeniile Vânju Mare":      "Entrepreneur / Owner",
 "Licorna WineHouse":         "Entrepreneur / Owner",
 "Gitana Winery":             "Entrepreneur / Owner",
 "Domeniile Sâmburești":      "Entrepreneur / Owner",
 "Crama Rasova":              "Entrepreneur / Owner",
 "Liliac Winery":             "Marketing / Sales",
 "Crama Oprișor":             "General Management",
 "Domeniile Săhăteni":        "Wine Engineer",
 "Crama 1000 de Chipuri":     "Entrepreneur / Owner",
 "Casa de Vinuri Negrini":    "Entrepreneur / Owner",
 "Petro Vaselo":              "Entrepreneur / Owner",
 "Carastelec Winery":         "Wine Engineer",
 "Domeniile Davidescu":       "Entrepreneur / Owner",
 "SERVE Ceptura":             "Entrepreneur / Owner",
 "Sogrape":                   "General Management",
 "Bodega Pietroasa":          "Not established",
 "Cricova":                   "Not established",
 "Ostrovit":                  "Entrepreneur / Owner",
 "Rovinex":                   "Not established",
 "Vinarte":                   "Not established",
 "Vinexport":                 "Not established",
 "Carpatvin":                 "Not established",
 "Chateau Vartely":           "Wine Engineer",
 "Bucium":                    "Not established",
 "Domeniile Panciu":          "Entrepreneur / Owner",
 "Doina Vin":                 "Not established",
 "Rifco Import":              "Not established",
 "Corcova":                   "Entrepreneur / Owner",
 "DC Segarcea":               "Agronomy / Horticulture",
 "Casa Isarescu":             "Finance / Economics",
 "Castel Mimi":               "General Management",
 "Alexandrion":               "Entrepreneur / Owner",
 "Suvorov Vin":               "Not established",
 "Barefoot Cellars":          "Entrepreneur / Owner",
 "Domeniul Bogdan":           "Entrepreneur / Owner",
 "Crama Basilescu":           "Entrepreneur / Owner",
 "Avincis Vinuri":            "Agronomy / Horticulture",
 "Crama La Salina":           "Not established",
 "Mastegariu Florin":         "Not established",
 "Unicom":                    "Not established",
 "Antinori":                  "Entrepreneur / Owner",
 "Fautor":                    "Wine Engineer",
 "Caraprodvin":               "Not established",
 "Grand Tokaj":               "General Management",
 "Vinia":                     "General Management",
 "Domeniul Burcilor":         "Not established",
 "Prodimas":                  "Not established",
 "WineRo":                    "Not established",
 "Casa De Vinuri Ciumbrud":   "Entrepreneur / Owner",
 "Domeniile Blaga":           "Not established",
 "Carl Reh Winery":           "General Management",
 "Amb Wine Company":          "Marketing / Sales",
 "Vintruvian":                "Academia",
}

wb = openpyxl.load_workbook(F)
ws = wb["CEO Research"]
last = ws.max_row
widths = {c: ws.column_dimensions[get_column_letter(c)].width for c in range(1, ws.max_column + 1)}

ws.insert_cols(COL)

hdr_style = copy(ws.cell(row=1, column=COL - 1)._style)
body_style = copy(ws.cell(row=3, column=COL - 1)._style)
last_style = copy(ws.cell(row=last, column=COL - 1)._style)

ws.cell(row=1, column=COL).value = HEADER
ws.cell(row=1, column=COL)._style = hdr_style

missing = []
for r in range(2, last + 1):
    winery = ws.cell(row=r, column=1).value
    val = CORE.get(winery)
    if val is None:
        missing.append(winery)
    cell = ws.cell(row=r, column=COL)
    cell.value = val
    cell._style = copy(body_style)
    cell.alignment = Alignment(wrap_text=True, vertical="top")
ws.cell(row=last, column=COL)._style = copy(last_style)
ws.cell(row=last, column=COL).alignment = Alignment(wrap_text=True, vertical="top")
if missing:
    raise SystemExit(f"no core discipline mapped for: {missing}")

# restore the original widths either side of the insert, and size the new one
for c, w in widths.items():
    if w:
        target = c if c < COL else c + 1
        ws.column_dimensions[get_column_letter(target)].width = w
ws.column_dimensions[get_column_letter(COL)].width = 21

dv = DataValidation(type="list", formula1='"' + ",".join(VOCAB) + '"',
                    allow_blank=False, showDropDown=False)
dv.error = "Use one of the eleven values in the legend at the foot of this sheet."
dv.errorTitle = "Core discipline"
dv.prompt = "Pick one standardized value."
dv.promptTitle = "Core discipline"
ws.add_data_validation(dv)
dv.add(f"{get_column_letter(COL)}2:{get_column_letter(COL)}{last}")

note = last + 2
ws.merge_cells(start_row=note, start_column=1, end_row=note, end_column=ws.max_column)
nc = ws.cell(row=note, column=1)
nc.value = (
    "Core discipline — one standardized value per leader: the profession they trained in or built their career on "
    "before wine, so the column can be filtered and counted. Where someone holds more than one, the documented "
    "education decides; where the education is not documented, the dominant career domain does. Permitted values: "
    + " · ".join(VOCAB) + ". "
    "\"Entrepreneur / Owner\" means a self-made proprietor with no single prior profession on the record — it is a "
    "finding, not a gap. \"Not established\" means no leader was identified, or none of their background is public — "
    "that is a gap. \"Wine Engineer\" is reserved for documented oenology or wine-technology training; running a "
    "winery is not by itself enough. The unstandardized detail stays in the Background, Education and Career domains "
    "columns; this column never replaces them.")
nc.font = Font(name="Arial", size=8, color="FF666666")
nc.alignment = Alignment(wrap_text=True, vertical="top")
ws.row_dimensions[note].height = 58

wb.save(F)

import collections
dist = collections.Counter(CORE[ws.cell(row=r, column=1).value] for r in range(2, last + 1))
print(f"CEO Research: column {get_column_letter(COL)} added, {last-1} rows classified, legend at row {note}")
for k, n in dist.most_common():
    print(f"  {k:26} {n}")
