# -*- coding: utf-8 -*-
"""Append the 61-company leadership research to CEO Research and Commercial People."""
from copy import copy
import openpyxl
from openpyxl.utils import get_column_letter
from companies import CEO_ROWS, PEOPLE_ROWS

F="Romanian_Wineries_Marketing_finalfinal.xlsx"
wb=openpyxl.load_workbook(F)

def append(ws,ncols,rows):
    old=ws.max_row
    st_mid=[copy(ws.cell(row=3,column=c)._style) for c in range(1,ncols+1)]
    st_last=[copy(ws.cell(row=old,column=c)._style) for c in range(1,ncols+1)]
    for c in range(1,ncols+1): ws.cell(row=old,column=c)._style=copy(st_mid[c-1])
    for i,row in enumerate(rows):
        r=old+1+i
        for c in range(1,ncols+1):
            ws.cell(row=r,column=c).value=row[c-1]
            ws.cell(row=r,column=c)._style=copy(st_mid[c-1])
    new=old+len(rows)
    for c in range(1,ncols+1): ws.cell(row=new,column=c)._style=copy(st_last[c-1])
    return old,new

ceo=wb["CEO Research"]
o,n=append(ceo,8,CEO_ROWS); print(f"CEO Research      {o} -> {n}")
cp=wb["Commercial People"]
cp_old,cp_last=append(cp,7,PEOPLE_ROWS); print(f"Commercial People {cp_old} -> {cp_last}")

def widen(ws,coords,pairs):
    for coord in coords:
        v=ws[coord].value
        if isinstance(v,str) and v.startswith("="):
            for a,b in pairs: v=v.replace(a,b)
            ws[coord]=v
pr=[(f"$C$2:$C${cp_old}",f"$C$2:$C${cp_last}"),(f"$E$2:$E${cp_old}",f"$E$2:$E${cp_last}")]
widen(cp,["J2","J3"]+[f"J{r}" for r in range(7,14)]+[f"M{r}" for r in range(7,18)],pr)

ts=wb["Commercial Type Split"]
TOTAL=33; NOTE=35                     # current layout: data 3-32, total 33, note 35
assert ts.cell(row=TOTAL,column=1).value=="Company total"
note_text=ts.cell(row=NOTE,column=1).value
note_style=copy(ts.cell(row=NOTE,column=1)._style)
ts.unmerge_cells(f"A{NOTE}:V{NOTE}")
for c in range(1,23): ts.cell(row=NOTE,column=c).value=None

st_data=[copy(ts.cell(row=3,column=c)._style) for c in range(1,23)]
st_dlast=[copy(ts.cell(row=TOTAL-1,column=c)._style) for c in range(1,23)]
st_total=[copy(ts.cell(row=TOTAL,column=c)._style) for c in range(1,23)]
tot_vals=[ts.cell(row=TOTAL,column=c).value for c in range(1,23)]
for c in range(1,23):
    ts.cell(row=TOTAL,column=c).value=None
    ts.cell(row=TOTAL,column=c)._style=copy(st_data[c-1])
    ts.cell(row=TOTAL-1,column=c)._style=copy(st_data[c-1])

tsp=[(f"$A$2:$A${cp_old}",f"$A$2:$A${cp_last}"),(f"$C$2:$C${cp_old}",f"$C$2:$C${cp_last}"),
     (f"$E$2:$E${cp_old}",f"$E$2:$E${cp_last}")]
widen(ts,[f"{get_column_letter(c)}{r}" for r in range(3,TOTAL)
          for c in list(range(2,9))+list(range(10,21))],tsp)

new_cos=sorted({p[0] for p in PEOPLE_ROWS})
for i,name in enumerate(new_cos):
    r=TOTAL+i
    ts.cell(row=r,column=1).value=name
    for c in range(2,9):
        col=get_column_letter(c)
        ts.cell(row=r,column=c).value=(f"=COUNTIFS('Commercial People'!$A$2:$A${cp_last},\"{name}\","
          f"'Commercial People'!$C$2:$C${cp_last},\"Marketing\",'Commercial People'!$E$2:$E${cp_last},{col}$2)")
    ts.cell(row=r,column=9).value=f"=SUM(B{r}:H{r})"
    for c in range(10,21):
        col=get_column_letter(c)
        ts.cell(row=r,column=c).value=(f"=COUNTIFS('Commercial People'!$A$2:$A${cp_last},\"{name}\","
          f"'Commercial People'!$C$2:$C${cp_last},\"Sales\",'Commercial People'!$E$2:$E${cp_last},{col}$2)")
    ts.cell(row=r,column=21).value=f"=SUM(J{r}:T{r})"
    ts.cell(row=r,column=22).value=f"=I{r}+U{r}"
    for c in range(1,23): ts.cell(row=r,column=c)._style=copy(st_data[c-1])
    ts.row_dimensions[r].height=22.5
dlast=TOTAL+len(new_cos)-1
for c in range(1,23): ts.cell(row=dlast,column=c)._style=copy(st_dlast[c-1])
nt=dlast+1
for c in range(1,23):
    v=tot_vals[c-1]
    if isinstance(v,str) and v.startswith("=SUM("):
        col=get_column_letter(c); v=f"=SUM({col}3:{col}{dlast})"
    ts.cell(row=nt,column=c).value=v
    ts.cell(row=nt,column=c)._style=copy(st_total[c-1])
ts.row_dimensions[nt].height=22.5
nn=nt+2
ts.merge_cells(start_row=nn,start_column=1,end_row=nn,end_column=22)
ts.cell(row=nn,column=1).value=note_text
ts.cell(row=nn,column=1)._style=copy(note_style)
ts.row_dimensions[nn].height=30.0
print(f"Type Split: data 3-{dlast}, total {nt}, note {nn}")
wb.save(F)
