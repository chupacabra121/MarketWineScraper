"""Add the Crama Ceptura research into Romanian_Wineries_Marketing_finalfinal.xlsx.

Rows are inserted next to the winery's existing block on each sheet so the
grouping is kept, styles are carried over from the surrounding rows, and every
COUNTIF/COUNTIFS range that pointed at the old last row is widened.
"""
import shutil
from copy import copy

import openpyxl

SRC = "original.xlsx"
OUT = "Romanian_Wineries_Marketing_finalfinal.xlsx"

shutil.copy(SRC, OUT)
wb = openpyxl.load_workbook(OUT)


# ---------------------------------------------------------------- helpers ---
def grab_style(ws, row, ncols):
    return [copy(ws.cell(row=row, column=c)._style) for c in range(1, ncols + 1)]


def put_style(ws, row, style, ncols):
    for c in range(1, ncols + 1):
        ws.cell(row=row, column=c)._style = copy(style[c - 1])


def rewrite_block(ws, ncols, new_rows, after_value, match_col=1):
    """Read rows 2..max, splice new_rows in after the last row whose match_col
    equals after_value, and write the result back with the sheet's own styles."""
    old_last = ws.max_row
    data = [
        [ws.cell(row=r, column=c).value for c in range(1, ncols + 1)]
        for r in range(2, old_last + 1)
    ]

    style_first = grab_style(ws, 2, ncols)
    style_mid = grab_style(ws, (2 + old_last) // 2, ncols)
    style_last = grab_style(ws, old_last, ncols)

    idx = max(i for i, row in enumerate(data) if row[match_col - 1] == after_value)
    data[idx + 1: idx + 1] = new_rows

    new_last = 1 + len(data)
    for i, row in enumerate(data):
        r = 2 + i
        for c in range(1, ncols + 1):
            ws.cell(row=r, column=c).value = row[c - 1]
        put_style(ws, r, style_mid, ncols)
    put_style(ws, 2, style_first, ncols)
    put_style(ws, new_last, style_last, ncols)
    return old_last, new_last


def widen(ws, cells, pairs):
    for coord in cells:
        v = ws[coord].value
        if isinstance(v, str) and v.startswith("="):
            for old, new in pairs:
                v = v.replace(old, new)
            ws[coord] = v


CHECKED = "Public sources checked 2026-08-21"

# ------------------------------------------------------- 1. Open Positions ---
op_rows = [
    [
        "Crama Ceptura",
        "IT Help Desk (Administrator de rețea)",
        "Recent vacancy",
        "Comuna Ceptura, Prahova",
        "Closed / not actively recruiting",
        "BestJobs listing checked 2026-08-21",
        "Help-desk role advertised by the winery itself on BestJobs: IT degree, 1–3 years "
        "support experience, Windows/Linux, Active Directory and Microsoft 365. The listing "
        "states recruitment is no longer active but still accepts expressions of interest.",
        "https://www.bestjobs.eu/loc-de-munca/administrator-de-retea-50",
    ],
    [
        "Crama Ceptura",
        "BestJobs employer profile is the winery's public recruitment route",
        "Job-board review",
        "Comuna Ceptura, Prahova",
        "Employer profile active; no open role at check",
        "BestJobs company profile checked 2026-08-21",
        "Crama Ceptura recruits through a BestJobs employer profile (57 alumni listed) rather "
        "than a careers page of its own, which is why the winery and group sites show nothing. "
        "This is the applicant route to record for the Romanian entity.",
        "https://www.bestjobs.eu/company-profile/crama-ceptura",
    ],
]
op_old, op_new = rewrite_block(wb["Open Positions"], 8, op_rows, "Crama Ceptura")

# ---------------------------------------------------- 2. Campaign Evidence ---
ce_rows = [
    [
        "Crama Ceptura",
        "2019",
        "„Ne aduce împreună” brand repositioning (Roadtrip)",
        "Brand platform launch",
        "Repositioning TVC built on a road-trip story and a 1980s Romanian song, agency "
        "United Pencils; moved the brand off product-and-terroir messaging onto everyday "
        "social occasions",
        "The origin of the togetherness platform: identity is built on people and occasions "
        "rather than on the vineyard, which is what lets one masterbrand carry several price tiers.",
        "https://www.iqads.ro/articol/48744/case-study-crama-ceptura-o-campanie-fireasca-despre-oameni",
    ],
    [
        "Crama Ceptura",
        "Jul. 2020",
        "„Împreună pentru mai bine!”",
        "Solidarity / purpose",
        "Pandemic-era campaign calling on Romanian entrepreneurs to support one another",
        "Stretches the togetherness platform from consumer occasions to business solidarity, "
        "keeping one message live through a period with no events to sponsor.",
        "https://www.iqads.ro/articol/51170/impreuna-pentru-mai-bine-crama-ceptura-lanseaza-o-campanie-indemn-la",
    ],
    [
        "Crama Ceptura",
        "Nov. 2021",
        "Magic FM „Magic Hour” co-promotion",
        "Radio partnership",
        "Daily on-air contest over three weeks, 15 prizes pairing a 3-bottle Magnus set with "
        "a set of books, SMS entry",
        "Borrows a station's own audience rather than buying spots, and matches the prize "
        "(wine plus books) to that station's listener profile.",
        "https://www.magicfm.ro/pagini/178/",
    ],
    [
        "Crama Ceptura",
        "Dec. 2021",
        "„Împreună, dincolo de granițe”",
        "TV brand campaign",
        "Humour-led TV, online and social films about a mixed-nationality family meeting, "
        "agency KUBIS",
        "Renews the platform by testing it against a topical Romanian reality — families "
        "spread across borders — instead of restating it.",
        "https://www.iqads.ro/articol/57608/kubis-semneaza-noua-campanie-tv-pentru-crama-ceptura-impreuna-dincolo-de-granite",
    ],
    [
        "Crama Ceptura",
        "Nov. 2024",
        "Rock FM „Vinuri premiate”",
        "Radio partnership",
        "Five daily prizes of six Purcari-portfolio bottles (476 lei each), entry by SMS or "
        "WhatsApp voice message, organised by SC Crama Ceptura SRL",
        "Shows the Romanian entity running promotions for the whole group portfolio, not only "
        "its own labels — the local marketing team is the group's route to Romanian consumers.",
        "https://www.rockfm.ro/pagini/311/regulament-rock-fm-campanie-promotionala-vinuri-premiate",
    ],
    [
        "Crama Ceptura",
        "Feb. 2025",
        "Brand-origin dispute over the „Ceptura” name",
        "Reputation risk",
        "Trade press reported 2024 rosé sold under the registered Crama Ceptura mark whose "
        "back label declared wine produced in the Republic of Moldova, while Ceptura is a "
        "protected sub-appellation; a competing winery leader called it a breach of the wine law",
        "The cost of a place-name masterbrand: sourcing outside the appellation turns the "
        "brand's strongest asset into a line of attack for competitors and trade media.",
        "https://vinul.ro/purcari-crama-ceptura-incalcarea-legislatiei-vinului.html",
    ],
    [
        "Crama Ceptura",
        "Feb. 2025",
        "SIWS Expo and Vinexpo America, APEV Romania pavilion",
        "Trade / export",
        "Exhibited within the Romanian producers' association pavilion at domestic and US "
        "trade fairs",
        "Keeps a trade-facing calendar running alongside the consumer platform; this is how "
        "the distributor and export side is worked, separately from the TV and radio work.",
        "https://apev.info.ro/2025-02-siws-expo-crama-ceptura/",
    ],
    [
        "Crama Ceptura",
        "2026",
        "Vinalies Internationales 2026 medals",
        "Awards / reputation",
        "Three golds and one silver — Collection Motiv Rosé and Alb, Collection Magnus Rosé "
        "and Alb, all 2025 vintages — within Romania's 27 medals at the edition",
        "The medals land on the mainstream Motiv and Magnus ranges rather than on a flagship, "
        "so third-party validation reaches the volume tiers that actually carry the brand.",
        "https://winesofromania.com/vinalies-internationales-2026-medalii-romania/",
    ],
    [
        "Crama Ceptura",
        "2026",
        "Decanter World Wine Awards 2026",
        "Awards / reputation",
        "Six medals, level with Crama Gîrboiu and behind only Casa de Vinuri Cotnari and "
        "Murfatlar among Romanian producers, out of 112 Romanian medals",
        "Sustains the annual awards drumbeat begun with the 2024 Decanter cycle and gives the "
        "trade a fixed yearly comparison point against the other large producers.",
        "https://winesofromania.com/en/romania-decanter-world-wine-awards-2026/",
    ],
    [
        "Crama Ceptura",
        "May 2026",
        "SERVE Ceptura acquisition completed",
        "Corporate growth",
        "Purcari completed the purchase of SERVE Ceptura — Romania's first private winery "
        "after 1989, about 60 ha and 500,000 bottles a year — sited 1.5 km from Crama Ceptura",
        "Buys a heritage and founder story the mass-market masterbrand cannot tell about "
        "itself, and puts both Ceptura names under one owner.",
        "https://www.zf.ro/burse-fonduri-mutuale/bursa-purcari-a-finalizat-achizitia-cramei-serve-ceptura-una-dintre-23154069",
    ],
    [
        "Crama Ceptura",
        "Recurring",
        "„Cu 14 zile mai mult soare”",
        "Terroir claim",
        "Standing Dealu Mare positioning line carried across the winery's own site, trade "
        "profiles and producer write-ups",
        "One quantified terroir claim does the work an origin story would, and sits alongside "
        "the people-first masterbrand message without competing with it.",
        "https://usatradetasting.com/en/blog/producer-profiles-76/crama-ceptura-a-vineyard-with-14-days-of-extra-sunshine-381.htm",
    ],
]
ce_old, ce_new = rewrite_block(wb["Campaign Evidence"], 7, ce_rows, "Crama Ceptura")

# ----------------------------------------------------- 3. Commercial People ---
cp_rows = [
    [
        "Crama Ceptura",
        "Cătălin Velicu",
        "Sales",
        "Regional Sales Manager",
        "Regional / Area Sales",
        "Manager",
        "https://ro.linkedin.com/in/catalin-velicu-051259b9",
    ],
]
cp = wb["Commercial People"]
cp_old, cp_new = rewrite_block(cp, 7, cp_rows, "Purcari Wineries / Crama Ceptura")

cp_pairs = [("$C$2:$C$87", f"$C$2:$C${cp_new}"), ("$E$2:$E$87", f"$E$2:$E${cp_new}")]
widen(cp, ["J2", "J3"] + [f"J{r}" for r in range(7, 14)] + [f"M{r}" for r in range(7, 18)], cp_pairs)

# ------------------------------------------------- 4. Commercial Type Split ---
ts = wb["Commercial Type Split"]
ts_pairs = [
    ("$A$2:$A$87", f"$A$2:$A${cp_new}"),
    ("$C$2:$C$87", f"$C$2:$C${cp_new}"),
    ("$E$2:$E$87", f"$E$2:$E${cp_new}"),
]
widen(ts, [f"{col}{r}" for r in range(3, 18) for col in "BCDEFGHJKLMNOPQRST"], ts_pairs)

# Ceptura now has people filed under its own label as well as the shared one;
# count both, the way the Purcari, Recaș and Via Viticola rows already do.
for col in "BCDEFGH":
    base = ts[f"{col}8"].value
    ts[f"{col}8"] = base + base.replace("=", "+", 1).replace(
        '"Purcari Wineries / Crama Ceptura"', '"Crama Ceptura"'
    )
for col in "JKLMNOPQRST":
    base = ts[f"{col}8"].value
    ts[f"{col}8"] = base + base.replace("=", "+", 1).replace(
        '"Purcari Wineries / Crama Ceptura"', '"Crama Ceptura"'
    )

# ------------------------------------------------------ 5. Takeaways sheet ---
tk = wb["Takeaways_Insights"]
widen(
    tk,
    [f"D{r}" for r in range(2, tk.max_row + 1)],
    [("$A$2:$A$27", f"$A$2:$A${op_new}"), ("$C$2:$C$27", f"$C$2:$C${op_new}")],
)

cep = next(r for r in range(2, tk.max_row + 1) if tk.cell(row=r, column=1).value == "Crama Ceptura")
tk.cell(row=cep, column=5).value = (
    "No open Ceptura role at check. The winery recruits through a BestJobs employer profile "
    "(57 alumni listed) rather than a careers page, and its IT Help Desk listing there is "
    "closed to new applications."
)
tk.cell(row=cep, column=6).value = (
    "A people-first „Ne aduce împreună” masterbrand platform running since 2019, worked "
    "through TV, recurring radio-station co-promotions and an annual awards drumbeat"
)
tk.cell(row=cep, column=7).value = (
    "Ne aduce împreună (2019); Împreună pentru mai bine (2020); Împreună, dincolo de granițe "
    "(2021); Magic FM, Rock FM and Radio ZU co-promotions; Decanter and Vinalies 2026; "
    "SERVE Ceptura acquisition"
)
tk.cell(row=cep, column=9).value = (
    "Purcari subsidiary and the group's Romanian commercial entity: it organises promotions "
    "for the whole Purcari portfolio, so part of its commercial roster is shared with Purcari. "
    "LinkedIn shows five sales-function members but only one is publicly identifiable by name. "
    "The group has been about 73% Maspex-owned since 2025. A Feb. 2025 trade-press dispute "
    "over Moldovan-produced wine sold under the Ceptura mark is a live brand-origin risk. "
    f"Campaign, hiring and awards evidence refreshed 2026-08-21; LinkedIn counts unchanged "
    "from the 2026-08-18 capture."
)

# ---------------------------------------------------------- 6. CEO Research ---
ceo = wb["CEO Research"]
crow = next(r for r in range(2, ceo.max_row + 1) if ceo.cell(row=r, column=1).value == "Crama Ceptura")
ceo.cell(row=crow, column=8).value = (
    "High confidence from the group's official executive biography, independently confirmed "
    "by trade-press campaign credits naming him General Manager, Crama Ceptura (IQads, "
    "Dec. 2021). Verified again 2026-08-21."
)

wb.save(OUT)
print("written:", OUT)
print(f"Open Positions rows {op_old} -> {op_new}")
print(f"Campaign Evidence rows {ce_old} -> {ce_new}")
print(f"Commercial People rows {cp_old} -> {cp_new}")
