"""The study's Excel deliverable, in German or English.

Sheet order follows how the question gets answered rather than how the data was
collected: the headline price points first, then the cuts that explain them
(segment, size, retailer, competing formats), then the two sheets that say what
the numbers do *not* cover — the PET evidence and the retailers that could not be
reached — and only then the raw rows.

Every figure on every summary sheet is computed from the rows on the data sheet,
so the workbook can be checked against itself. Every user-visible string comes
from :mod:`.text`, so the German and English builds differ in wording and in
nothing else: same sheets, same rows, same numbers.
"""

from __future__ import annotations

import statistics
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Callable, Iterable, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from . import brands
from . import packaging as pkg
from .model import EXPORT_COLUMNS, Listing
from .sources import UNAVAILABLE
from .text import Texts

# --- house style -------------------------------------------------------------
INK = "1F2933"
ACCENT = "7A2E3C"          # wine red, used for headers
ACCENT_SOFT = "F2E8EA"
RULE = "D8DEE4"
MUTED = "6B7785"

TITLE_FONT = Font(name="Calibri", size=16, bold=True, color=ACCENT)
SUB_FONT = Font(name="Calibri", size=10, color=MUTED, italic=True)
HEAD_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
BODY_FONT = Font(name="Calibri", size=10, color=INK)
STRONG = Font(name="Calibri", size=10, bold=True, color=INK)
NOTE_FONT = Font(name="Calibri", size=9, color=MUTED)

HEAD_FILL = PatternFill("solid", fgColor=ACCENT)
BAND_FILL = PatternFill("solid", fgColor=ACCENT_SOFT)
THIN = Side(style="thin", color=RULE)
CELL_BORDER = Border(bottom=THIN)

EUR = '#,##0.00\\ "€"'
EUR_L = '#,##0.00\\ "€/l"'
NUM2 = "0.00"
PCT = "0%"


def _autosize(sheet: Worksheet, widths: Sequence[int]) -> None:
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width


def _title(sheet: Worksheet, title: str, subtitle: str = "", *, span: int = 8) -> int:
    """Write the sheet's title block and return the next free row."""
    sheet["A1"] = title
    sheet["A1"].font = TITLE_FONT
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    row = 2
    if subtitle:
        sheet.cell(row=2, column=1, value=subtitle).font = SUB_FONT
        sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=span)
        row = 3
    sheet.row_dimensions[1].height = 24
    return row + 1


def _table(sheet: Worksheet, start_row: int, headers: Sequence[str],
           rows: Iterable[Sequence[Any]], formats: dict[int, str] | None = None,
           *, band: bool = True) -> int:
    """Write a header row and body, and return the row after the table."""
    formats = formats or {}
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(row=start_row, column=column, value=header)
        cell.font = HEAD_FONT
        cell.fill = HEAD_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    sheet.row_dimensions[start_row].height = 28

    row_number = start_row
    for offset, values in enumerate(rows):
        row_number = start_row + 1 + offset
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row=row_number, column=column, value=value)
            cell.font = BODY_FONT
            cell.border = CELL_BORDER
            if band and offset % 2 == 1:
                cell.fill = BAND_FILL
            if column in formats:
                cell.number_format = formats[column]
    return row_number + 2


def _note(sheet: Worksheet, row: int, text: str, *, span: int = 8) -> int:
    cell = sheet.cell(row=row, column=1, value=text)
    cell.font = NOTE_FONT
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    sheet.row_dimensions[row].height = max(14, 13 * (len(text) // 110 + 1))
    return row + 2


# --- statistics --------------------------------------------------------------
def _stats(values: list[float]) -> dict[str, float | None]:
    if not values:
        return {"n": 0, "min": None, "median": None, "mean": None, "max": None}
    ordered = sorted(values)
    return {
        "n": len(ordered),
        "min": round(ordered[0], 2),
        "median": round(statistics.median(ordered), 2),
        "mean": round(statistics.fmean(ordered), 2),
        "max": round(ordered[-1], 2),
    }


def _by(listings: list[Listing], key: Callable[[Listing], Any]) -> dict[Any, list[Listing]]:
    grouped: dict[Any, list[Listing]] = {}
    for listing in listings:
        grouped.setdefault(key(listing), []).append(listing)
    return grouped


def _litre_prices(listings: Iterable[Listing]) -> list[float]:
    return [x.price_per_litre for x in listings if x.price_per_litre is not None]


def _money(value: float, t: Texts) -> str:
    """A euro amount written the way the language writes it."""
    if t.language == "de":
        return f"{value:.2f}".replace(".", ",") + " €"
    return f"{value:.2f} €"


# --- sheets ------------------------------------------------------------------
def _sheet_summary(book: Workbook, scope: list[Listing], everything: list[Listing],
                   stamp: str, t: Texts) -> None:
    sheet = book.create_sheet(t("sheet_summary"))
    row = _title(sheet, t("summary_title"), t("summary_sub", stamp=stamp))

    consumer = [x for x in scope if x.price_basis == "gross"]
    still = [x for x in consumer if x.product_type == "still"]

    row = _note(sheet, row, t("summary_headline"))

    litre = _litre_prices(still)
    boxes = [x for x in still if x.packaging == pkg.BAG_IN_BOX]
    three = [x for x in boxes if x.volume_l == 3.0]
    facts = [
        (t("fact_total"), len(everything), t("fact_total_note")),
        (t("fact_scope"), len(scope),
         t("fact_scope_note", pct=f"{len(scope)/len(everything):.0%}")),
        (t("fact_bib"), sum(1 for x in scope if x.packaging == pkg.BAG_IN_BOX), ""),
        (t("fact_pet"), sum(1 for x in scope if x.packaging == pkg.PET),
         t("fact_pet_note")),
        (t("fact_still"), len(still), t("fact_still_note")),
    ]
    row = _table(sheet, row, [t("h_metric"), t("h_value"), t("h_note")],
                 facts, {2: "#,##0"})

    stats = _stats(litre)
    box_stats = _stats(_litre_prices(boxes))
    three_stats = _stats(_litre_prices(three))
    price_rows = [
        (t("row_all_scope"), stats["n"], stats["min"], stats["median"],
         stats["mean"], stats["max"]),
        (t("row_bib_only"), box_stats["n"], box_stats["min"], box_stats["median"],
         box_stats["mean"], box_stats["max"]),
        (t("row_three_only"), three_stats["n"], three_stats["min"],
         three_stats["median"], three_stats["mean"], three_stats["max"]),
    ]
    row = _table(sheet, row,
                 [t("h_price_point"), t("h_offers"), t("h_min"), t("h_median"),
                  t("h_mean"), t("h_max")],
                 price_rows, {3: EUR_L, 4: EUR_L, 5: EUR_L, 6: EUR_L})

    # The 0.75 L equivalent is the comparison a shopper actually makes.
    equivalents = [x.bottle_equivalent_price for x in three
                   if x.bottle_equivalent_price is not None]
    eq = _stats(equivalents)
    if eq["n"]:
        row = _table(sheet, row,
                     [t("h_bottle_equiv"), t("h_offers"), t("h_min"),
                      t("h_median"), t("h_mean"), t("h_max")],
                     [(t("row_bottle_equiv"), eq["n"], eq["min"], eq["median"],
                       eq["mean"], eq["max"])],
                     {3: EUR, 4: EUR, 5: EUR, 6: EUR})

    row = _note(sheet, row, t("note_pfand"))
    _note(sheet, row, t("note_metro"))
    _autosize(sheet, [38, 12, 14, 12, 13, 12, 12, 12])


def _sheet_segments(book: Workbook, scope: list[Listing], t: Texts) -> None:
    sheet = book.create_sheet(t("sheet_segments"))
    row = _title(sheet, t("segments_title"), t("segments_sub"))

    three = [x for x in scope
             if x.packaging == pkg.BAG_IN_BOX and x.volume_l == 3.0
             and x.product_type == "still" and x.price_basis == "gross"
             and x.price is not None]

    # Bands cut at the gaps in the observed distribution, not at round numbers.
    # The floor is 4.99; the next cluster sits at 7.12; the median is 11.49 and
    # the upper quartile 15.99. So 8, 12 and 18 fall between groups rather than
    # through the middle of one.
    bands = [
        ("seg_entry", 0.0, 8.0, "seg_entry_note"),
        ("seg_mid", 8.0, 12.0, "seg_mid_note"),
        ("seg_premium", 12.0, 18.0, "seg_premium_note"),
        ("seg_top", 18.0, 10_000.0, "seg_top_note"),
    ]
    rows = []
    for name_key, low, high, note_key in bands:
        group = [x for x in three if low <= x.price < high]
        stats = _stats(_litre_prices(group))
        # The outer bands are open-ended and read better said that way than as
        # "0.00 €–8.00 €", which implies a floor nobody prices against.
        if low == 0.0:
            band = t("band_under", high=_money(high, t).replace(" €", ""))
        elif high >= 1000:
            band = t("band_from", low=_money(low, t).replace(" €", ""))
        else:
            band = f"{_money(low, t)}–{_money(high, t)}"
        rows.append((
            t(name_key), band, len(group),
            round(len(group) / len(three), 4) if three else 0,
            stats["min"], stats["median"], stats["max"], t(note_key),
        ))
    row = _table(sheet, row,
                 [t("h_segment"), t("h_price_band"), t("h_offers"), t("h_share"),
                  t("h_litre_min"), t("h_litre_med"), t("h_litre_max"),
                  t("h_typical")],
                 rows, {4: PCT, 5: EUR_L, 6: EUR_L, 7: EUR_L})

    row = _note(sheet, row, t("segments_note"), span=8)

    by_colour = []
    for colour, group in sorted(_by([x for x in three if x.colour],
                                    lambda x: x.colour).items()):
        stats = _stats(_litre_prices(group))
        by_colour.append((t.colour(colour), stats["n"], stats["min"],
                          stats["median"], stats["max"]))
    row = _table(sheet, row,
                 [t("h_colour_three"), t("h_offers"), t("h_litre_min"),
                  t("h_litre_med"), t("h_litre_max")],
                 by_colour, {3: EUR_L, 4: EUR_L, 5: EUR_L})

    by_country = []
    for country, group in sorted(_by([x for x in three if x.country],
                                     lambda x: x.country).items(),
                                 key=lambda kv: -len(kv[1])):
        stats = _stats(_litre_prices(group))
        by_country.append((t.country(country), stats["n"], stats["min"],
                           stats["median"], stats["max"]))
    _table(sheet, row,
           [t("h_origin_three"), t("h_offers"), t("h_litre_min"),
            t("h_litre_med"), t("h_litre_max")],
           by_country, {3: EUR_L, 4: EUR_L, 5: EUR_L})
    _autosize(sheet, [18, 20, 11, 10, 12, 14, 12, 42])


def _sheet_by_format(book: Workbook, scope: list[Listing], t: Texts) -> None:
    sheet = book.create_sheet(t("sheet_sizes"))
    row = _title(sheet, t("sizes_title"), t("sizes_sub"))

    still = [x for x in scope
             if x.product_type == "still" and x.price_basis == "gross"]
    # Single containers only. A "BiB pack, 9 L" is three boxes sold together,
    # and listing it as a nine-litre container would invent a size nobody fills.
    singles = [x for x in still if not x.is_pack]

    def _size_rows(group_source: list[Listing]) -> list[tuple]:
        out = []
        for volume, group in sorted(_by(group_source, lambda x: x.volume_l).items(),
                                    key=lambda kv: (kv[0] is None, kv[0])):
            stats = _stats(_litre_prices(group))
            prices = sorted(x.price for x in group if x.price is not None)
            out.append((
                f"{volume:g} l" if volume else t("not_stated"),
                t.packaging(group[0].packaging),
                # len(group), not stats["n"]: a row with no size has no price
                # per litre, and showing its offer count as zero would read as
                # an empty row rather than as "size not stated".
                len(group), prices[0] if prices else None,
                statistics.median(prices) if prices else None,
                prices[-1] if prices else None,
                stats["min"], stats["median"], stats["max"],
            ))
        return out

    headers = [t("h_container"), t("h_packaging"), t("h_offers"),
               t("h_price_min"), t("h_price_med"), t("h_price_max"),
               t("h_litre_min"), t("h_litre_med"), t("h_litre_max")]
    money = {4: EUR, 5: EUR, 6: EUR, 7: EUR_L, 8: EUR_L, 9: EUR_L}
    row = _table(sheet, row, headers, _size_rows(singles), money)
    row = _note(sheet, row, t("sizes_note"), span=9)

    packs = [x for x in still if x.is_pack]
    if packs:
        row = _table(sheet, row, [t("h_multipacks")] + headers[1:],
                     _size_rows(packs), money)
        _note(sheet, row, t("multipack_note"), span=9)
    _autosize(sheet, [34, 18, 11, 12, 14, 12, 12, 14, 12])


def _sheet_by_retailer(book: Workbook, scope: list[Listing], t: Texts) -> None:
    sheet = book.create_sheet(t("sheet_retailers"))
    row = _title(sheet, t("retailers_title"), t("retailers_sub"))

    rows = []
    for label, group in sorted(_by(scope, lambda x: x.retailer_label).items()):
        stats = _stats(_litre_prices(group))
        prices = sorted(x.price for x in group if x.price is not None)
        rows.append((
            label, t.channel(group[0].channel), t.basis(group[0].price_basis),
            stats["n"], prices[0] if prices else None,
            prices[-1] if prices else None,
            stats["min"], stats["median"], stats["max"],
        ))
    row = _table(sheet, row,
                 [t("h_retailer"), t("h_channel"), t("h_price_basis"),
                  t("h_offers"), t("h_price_min"), t("h_price_max"),
                  t("h_litre_min"), t("h_litre_med"), t("h_litre_max")],
                 rows, {5: EUR, 6: EUR, 7: EUR_L, 8: EUR_L, 9: EUR_L})

    _note(sheet, row, t("retailers_note"), span=9)
    _autosize(sheet, [24, 18, 14, 11, 12, 12, 12, 14, 12])


#: Why a cheaper listing was kept out of the per-store ranking. Keyed by the
#: product type that caused it, plus the pack case.
_EXCLUSION_KEYS = {
    "gluehwein": "excl_gluehwein", "sangria": "excl_sangria",
    "sparkling": "excl_sparkling", "dessert": "excl_dessert",
}


def _rank_key(listing: Listing) -> tuple:
    """Sort order for the cheapest-first ranking.

    Ties on price per litre are the normal case rather than the exception —
    Lidl prices three colours of its own-brand box identically, and Schäpers
    five — so the tie-break has to be deterministic or the sheet reshuffles
    between runs for no reason. Cheapest per litre first, then the smaller pack
    price, then the name.
    """
    return (listing.price_per_litre, listing.price or 0.0, listing.name)


def _sheet_cheapest(book: Workbook, scope: list[Listing], t: Texts,
                    per_store: int = 3) -> None:
    """The three cheapest bag-in-box wines at each store, per litre.

    Three filters are applied and each is stated rather than assumed, because
    all three change the answer at some store:

    * **Still wine only.** Glühwein sells in the same 10-litre box at a third of
      the litre price, and would take first place at METRO and WirWinzer.
    * **Single boxes only.** A "4er Paket" has an honest price per litre and is
      still not a thing a shopper can buy one of.
    * **METRO stays but is marked.** Its prices are net of VAT, so its 1.42 €/l
      is not the same kind of number as Lidl's 1.66.

    Whatever a filter removes that would have ranked is listed underneath with
    the reason, so the sheet cannot quietly flatter a store.
    """
    sheet = book.create_sheet(t("sheet_cheapest"))
    row = _title(sheet, t("cheapest_title"), t("cheapest_sub"), span=9)
    row = _note(sheet, row, t("cheapest_intro"), span=9)

    boxes = [x for x in scope
             if x.packaging == pkg.BAG_IN_BOX and x.price_per_litre is not None]
    # Identity, not equality: Listing compares by value, and two listings of the
    # same wine at the same price in different shops are equal without being the
    # same row.
    eligible = {id(x) for x in boxes
                if x.product_type == "still" and not x.is_pack}

    rows: list[tuple] = []
    excluded: list[tuple] = []
    ranked_pairs: list[tuple[str, Listing]] = []
    for label, group in sorted(_by(boxes, lambda x: x.retailer_label).items()):
        ranked = sorted([x for x in group if id(x) in eligible], key=_rank_key)
        top = ranked[:per_store]
        ranked_pairs.extend((label, x) for x in top)
        for position, listing in enumerate(top, start=1):
            rows.append((
                label if position == 1 else "", position,
                listing.name,
                f"{listing.volume_l:g} l" if listing.volume_l else t("not_stated"),
                listing.price, listing.price_per_litre,
                listing.bottle_equivalent_price,
                t.colour(listing.colour) if listing.colour else "",
                t.country(listing.country) if listing.country else "",
            ))
        if not top:
            rows.append((label, 1, t("cheapest_none"), "", None, None, None, "", ""))

        # Anything cheaper than the third-placed wine that a filter removed.
        ceiling = top[-1].price_per_litre if top else float("inf")
        for listing in sorted(group, key=_rank_key):
            if id(listing) in eligible or listing.price_per_litre > ceiling:
                continue
            reason = (t("excl_pack") if listing.is_pack
                      else t(_EXCLUSION_KEYS.get(listing.product_type,
                                                 "excl_gluehwein")))
            excluded.append((label, listing.name,
                             f"{listing.volume_l:g} l" if listing.volume_l else "",
                             listing.price, listing.price_per_litre, reason))

    row = _table(sheet, row,
                 [t("h_retailer"), t("h_rank"), t("h_wine"), t("h_size"),
                  t("h_price"), t("h_per_litre"), t("h_per_bottle"),
                  t("h_colour"), t("h_origin")],
                 rows, {5: EUR, 6: EUR_L, 7: EUR})

    row = _note(sheet, row, t("cheapest_note"), span=9)

    if excluded:
        _table(sheet, row,
               [t("h_retailer"), t("h_excluded"), t("h_size"), t("h_price"),
                t("h_per_litre"), t("h_why")],
               excluded, {4: EUR, 5: EUR_L})
    _autosize(sheet, [24, 7, 58, 10, 11, 12, 11, 10, 14])
    return [(label, listing) for label, listing in ranked_pairs]


def _sheet_private_label(book: Workbook, ranked: list[tuple[str, Listing]],
                         t: Texts) -> None:
    """Whether each ranked wine is the retailer's own label, and the source.

    The judgements live in :mod:`.brands`, hand-collected and committed, so this
    sheet only joins them to the ranking and renders the links. Where the
    sources did not settle a case the sheet prints "not established" rather than
    filling the gap — six of the twenty-seven end that way, and a confident
    guess in those rows would be the one thing a reader could not check.
    """
    sheet = book.create_sheet(t("sheet_label"))
    row = _title(sheet, t("label_title"), t("label_sub"), span=7)
    row = _note(sheet, row, t("label_method"), span=7)
    row = _note(sheet, row, t("label_finding"), span=7)

    verdict = {True: t("yes"), False: t("no"), None: t("unresolved")}
    header_row = row
    _table(sheet, row,
           [t("h_retailer"), t("h_wine"), t("h_private_label"),
            t("h_brand_owner"), t("h_operator"), t("h_basis"), t("h_sources")],
           [], band=False)
    row += 1

    for label, listing in ranked:
        record = brands.lookup(listing.retailer, listing.name)
        values = [
            label, listing.name,
            verdict[record.private_label] if record else t("unresolved"),
            record.brand_owner if record else "",
            record.operator if record else "",
            record.basis if record else "",
        ]
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row=row, column=column, value=value)
            cell.font = BODY_FONT
            cell.border = CELL_BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if column == 3 and record is not None:
                cell.font = STRONG
        # Each source as its own clickable cell, so a reader can follow any one
        # of them rather than picking a URL out of a joined string.
        for offset, url in enumerate((record.sources if record else [])):
            cell = sheet.cell(row=row, column=7 + offset, value=url)
            cell.hyperlink = url
            cell.font = Font(name="Calibri", size=9, color="0563C1",
                             underline="single")
            cell.border = CELL_BORDER
        sheet.row_dimensions[row].height = 46
        row += 1

    sheet.freeze_panes = sheet.cell(row=header_row + 1, column=1)
    _autosize(sheet, [22, 46, 13, 30, 38, 62, 46, 46, 46])


def _sheet_competing_formats(book: Workbook, everything: list[Listing],
                             t: Texts) -> None:
    """What bag-in-box is priced against on the same shelf."""
    sheet = book.create_sheet(t("sheet_formats"))
    row = _title(sheet, t("formats_title"), t("formats_sub"))

    consumer = [x for x in everything
                if x.price_basis == "gross" and x.product_type == "still"]
    rows = []
    for container, group in sorted(_by(consumer, lambda x: x.packaging).items(),
                                   key=lambda kv: -len(kv[1])):
        stats = _stats(_litre_prices(group))
        deposits = {x.pfand for x in group if x.pfand is not None}
        if deposits == {pkg.AMOUNT}:
            deposit = _money(pkg.AMOUNT, t)
        elif deposits == {0.0}:
            deposit = t("deposit_free")
        else:
            deposit = t("deposit_mixed")
        rows.append((t.packaging(container), stats["n"], stats["min"],
                     stats["median"], stats["max"], deposit))
    row = _table(sheet, row,
                 [t("h_packaging"), t("h_offers"), t("h_litre_min"),
                  t("h_litre_med"), t("h_litre_max"), t("h_deposit")],
                 rows, {3: EUR_L, 4: EUR_L, 5: EUR_L})

    row = _note(sheet, row, t("formats_note"), span=6)

    # The one comparison that decides whether the format is cheap: a 3-litre box
    # against the litre bottle and the carton, which sit in the same aisle.
    compare = []
    for key, group in (
        ("cmp_box", [x for x in consumer
                     if x.packaging == pkg.BAG_IN_BOX and x.volume_l == 3.0]),
        ("cmp_litre", [x for x in consumer if x.volume_l == 1.0]),
        ("cmp_carton", [x for x in consumer if x.packaging == pkg.CARTON]),
    ):
        stats = _stats(_litre_prices(group))
        compare.append((t(key), stats["n"], stats["min"], stats["median"],
                        stats["max"]))
    _table(sheet, row,
           [t("h_direct_compare"), t("h_offers"), t("h_litre_min"),
            t("h_litre_med"), t("h_litre_max")],
           compare, {3: EUR_L, 4: EUR_L, 5: EUR_L})
    _autosize(sheet, [32, 11, 12, 14, 12, 20])


def _sheet_pet(book: Workbook, probe_results: list, everything: list[Listing],
               t: Texts) -> None:
    sheet = book.create_sheet(t("sheet_pet"))
    row = _title(sheet, t("pet_title"), t("pet_sub"))
    row = _note(sheet, row, t("pet_intro"))

    # The strongest evidence is the census, not the search: 2,221 Globus wines
    # examined one by one is a firmer denominator than any keyword query.
    census = []
    for label, group in sorted(_by(everything, lambda x: x.retailer_label).items()):
        census.append((
            label, len(group),
            sum(1 for x in group if x.packaging == pkg.PET),
            sum(1 for x in group if x.packaging == pkg.BAG_IN_BOX),
        ))
    census.append((t("row_total"), len(everything),
                   sum(1 for x in everything if x.packaging == pkg.PET),
                   sum(1 for x in everything if x.packaging == pkg.BAG_IN_BOX)))
    row = _table(sheet, row,
                 [t("h_census"), t("h_checked"), t("h_of_pet"), t("h_of_bib")],
                 census, {2: "#,##0"})

    if probe_results:
        rows = [(r.source, r.query, r.hits, r.pet_hits, r.pet_wine_hits, r.example)
                for r in probe_results]
        row = _table(sheet, row,
                     [t("h_search"), t("h_query"), t("h_hits"), t("h_of_pet"),
                      t("h_pet_wine"), t("h_example")], rows)
        row = _note(sheet, row, t("pet_search_note"))

    context = [(t("pet_law"), t("pet_law_note")),
               (t("pet_supply"), t("pet_supply_note")),
               (t("pet_place"), t("pet_place_note")),
               (t("pet_where"), t("pet_where_note"))]
    _table(sheet, row, [t("h_point"), t("h_finding")], context)
    _autosize(sheet, [22, 30, 10, 11, 14, 46])


def _sheet_unavailable(book: Workbook, t: Texts) -> None:
    sheet = book.create_sheet(t("sheet_unavailable"))
    row = _title(sheet, t("unavailable_title"), t("unavailable_sub"), span=4)

    by_reason: dict[str, list[tuple]] = {}
    for _, label, channel, reason, detail in UNAVAILABLE:
        by_reason.setdefault(reason, []).append(
            (label, t.channel(channel), detail))

    # Grouped by *why*, because the reasons mean very different things and
    # lumping them together would read as one uniform failure.
    order = ["keine Preise online", "blockiert", "Preise clientseitig",
             "nicht erreichbar"]
    for reason in order + [r for r in by_reason if r not in order]:
        entries = by_reason.get(reason)
        if not entries:
            continue
        cell = sheet.cell(row=row, column=1,
                          value=f"{t.reason(reason)}  ({len(entries)})")
        cell.font = STRONG
        row += 1
        row = _table(sheet, row,
                     [t("h_retailer"), t("h_channel"), t("h_finding")],
                     sorted(entries))

    _note(sheet, row, t("unavailable_note"), span=4)
    _autosize(sheet, [30, 18, 104, 14])


def _sheet_data(book: Workbook, listings: list[Listing], title: str,
                sheet_name: str, t: Texts) -> None:
    sheet = book.create_sheet(sheet_name)
    # The column headers stay in English in both builds: they are the CSV field
    # names, and a reader matching the sheet against the exported file needs
    # them to be the same string.
    row = _title(sheet, title, t("data_sub"), span=len(EXPORT_COLUMNS))

    money = {EXPORT_COLUMNS.index(c) + 1: EUR
             for c in ("price", "pfand", "price_incl_pfand", "list_price",
                       "bottle_equivalent_price") if c in EXPORT_COLUMNS}
    money.update({EXPORT_COLUMNS.index(c) + 1: EUR_L
                  for c in ("price_per_litre", "price_per_litre_incl_pfand",
                            "unit_price") if c in EXPORT_COLUMNS})
    money[EXPORT_COLUMNS.index("volume_l") + 1] = NUM2

    body = []
    for listing in sorted(listings, key=lambda x: (x.retailer_label,
                                                   x.price_per_litre or 0)):
        flat = listing.to_row()
        body.append([flat.get(column) for column in EXPORT_COLUMNS])
    _table(sheet, row, EXPORT_COLUMNS, body, money, band=False)

    sheet.freeze_panes = sheet.cell(row=row + 1, column=1)
    sheet.auto_filter.ref = (f"A{row}:"
                             f"{get_column_letter(len(EXPORT_COLUMNS))}{row + len(body)}")
    widths = [14] * len(EXPORT_COLUMNS)
    for column in ("name", "url", "packaging_evidence", "category_path"):
        if column in EXPORT_COLUMNS:
            widths[EXPORT_COLUMNS.index(column)] = 52
    _autosize(sheet, widths)


def _sheet_method(book: Workbook, stamp: str, everything: list[Listing],
                  t: Texts) -> None:
    sheet = book.create_sheet(t("sheet_method"))
    row = _title(sheet, t("method_title"), t("method_sub", stamp=stamp))

    from .sources import all_sources
    rows = [(cls.label, t.channel(cls.channel), t.basis(cls.price_basis),
             sum(1 for x in everything if x.retailer == key), cls.note)
            for key, cls in all_sources().items()]
    row = _table(sheet, row,
                 [t("h_source"), t("h_channel"), t("h_price_basis"),
                  t("h_wines_collected"), t("h_access")], rows)

    checked = sum(1 for x in everything
                  if x.price_per_litre is not None and x.unit_price is not None)
    notes = [
        (t("m_scope"), t("m_scope_note")),
        (t("m_three_litre"), t("m_three_litre_note")),
        (t("m_per_litre"), t("m_per_litre_note")),
        (t("m_crosscheck"), t("m_crosscheck_note", checked=f"{checked:,}")),
        (t("m_packs"), t("m_packs_note")),
        (t("m_exclusions"), t("m_exclusions_note")),
        (t("m_limits"), t("m_limits_note")),
    ]
    _table(sheet, row, [t("h_point"), t("h_explanation")], notes)
    _autosize(sheet, [26, 110, 14, 14, 60])


# --- entry point -------------------------------------------------------------
def build_workbook(listings: list[Listing], path: Path,
                   probe_results: list | None = None,
                   language: str = "de") -> Path:
    """Write the study workbook in ``language`` and return where it went."""
    t = Texts(language)
    scope = [x for x in listings if pkg.is_in_scope(x.packaging)]
    now = datetime.now(timezone.utc)
    stamp = now.strftime("%d.%m.%Y" if language == "de" else "%d %B %Y")

    book = Workbook()
    book.remove(book.active)

    _sheet_summary(book, scope, listings, stamp, t)
    ranked = _sheet_cheapest(book, scope, t)
    _sheet_private_label(book, ranked, t)
    _sheet_segments(book, scope, t)
    _sheet_by_format(book, scope, t)
    _sheet_by_retailer(book, scope, t)
    _sheet_competing_formats(book, listings, t)
    _sheet_pet(book, probe_results or [], listings, t)
    _sheet_unavailable(book, t)
    _sheet_data(book, scope, t("data_title"), t("sheet_data"), t)
    _sheet_data(book, listings, t("all_title"), t("sheet_all"), t)
    _sheet_method(book, stamp, listings, t)

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    book.save(path)
    return path
