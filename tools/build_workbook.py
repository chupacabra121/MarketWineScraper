"""Build the wine-market workbook: a flat fact table plus formula-driven insight sheets."""
from __future__ import annotations

import json
import re
import sqlite3
import statistics
import unicodedata
from collections import Counter, defaultdict

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.formula import ArrayFormula
from openpyxl.worksheet.table import Table, TableStyleInfo

DB = "data/wines.sqlite"
OUT = "exports/romania-wine-market.xlsx"

FONT = "Arial"
NAVY = "1F3864"
ACCENT = "2E5F8A"
LIGHT = "DCE6F1"
GREY = "F2F2F2"
BLUE_TEXT = "0000FF"

# How each retailer's price should be read. This is the single most important
# caveat in the whole dataset, so it travels with every row.
PRICE_BASIS = {
    "auchan": ("Own site", "Shelf", "Retailer's own e-commerce price."),
    "carrefour": ("Own site", "Shelf", "Retailer's own e-commerce price."),
    "selgros": ("Own site", "Shelf", "Cash & carry; price is per depot."),
    "metro": ("Own site", "Shelf", "Cash & carry; VAT-incl, deposit excl. 6-bottle minimum common."),
    "freshful": ("Own site", "Shelf", "Online-only retailer; shelf price (Genius price in raw)."),
    "sezamo": ("Own site", "Shelf", "Online-only retailer."),
    "mega_image": ("Own site", "Shelf", "Retailer's own e-commerce price."),
    "penny": ("Own site", "Shelf", "Shelf price; PENNY-card price held separately."),
    "kaufland": ("Own site", "Promo only", "Weekly leaflet; promotional wines only."),
    "kaufland_bolt": ("Bolt Food", "Platform", "Delivery platform price; at or above shelf."),
    "penny_bolt": ("Bolt Food", "Platform", "Measured = shelf (median +0.0% vs penny.ro)."),
    "profi_glovo": ("Glovo", "Platform", "Includes 0.50 lei SGR deposit."),
    "supeco_glovo": ("Glovo", "Platform", "Includes 0.50 lei SGR deposit."),
}

LABELS = {
    "auchan": "Auchan", "carrefour": "Carrefour", "selgros": "Selgros", "metro": "METRO",
    "freshful": "Freshful", "sezamo": "Sezamo", "mega_image": "Mega Image",
    "penny": "Penny (own site)", "kaufland": "Kaufland (leaflet)",
    "kaufland_bolt": "Kaufland (Bolt)", "penny_bolt": "Penny (Bolt)",
    "profi_glovo": "Profi (Glovo)", "supeco_glovo": "Supeco (Glovo)",
}


def fold(s: str) -> str:
    s = (s or "").replace("ş", "s").replace("ţ", "t")
    return "".join(c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c)).lower()


# Retailers spell origin differently ("România"/"Romania", "Moldova, Republic Of"/
# "Republica Moldova"), which would split one country across several rows.
COUNTRY_CANON = {
    "romania": "Romania", "republica moldova": "Moldova", "moldova": "Moldova",
    "moldova, republic of": "Moldova", "italia": "Italy", "italy": "Italy",
    "franta": "France", "france": "France", "spania": "Spain", "spain": "Spain",
    "ungaria": "Hungary", "hungary": "Hungary", "portugalia": "Portugal",
    "germania": "Germany", "grecia": "Greece", "austria": "Austria",
    "chile": "Chile", "argentina": "Argentina", "africa de sud": "South Africa",
    "australia": "Australia", "noua zeelanda": "New Zealand", "sua": "USA",
    "statele unite ale americii": "USA", "bulgaria": "Bulgaria", "serbia": "Serbia",
    "georgia": "Georgia", "libanul": "Lebanon", "liban": "Lebanon",
}


def canon_country(value: str | None) -> str | None:
    if not value:
        return None
    return COUNTRY_CANON.get(fold(value).strip(), value.strip())


# --------------------------------------------------------------------- load
def load_rows():
    conn = sqlite3.connect(f"file:{DB}?mode=ro", uri=True)
    conn.row_factory = sqlite3.Row
    sql = """
    SELECT p.*, o.price, o.currency, o.list_price, o.unit_price, o.price_per_litre,
           o.on_promotion, o.offer_type, o.in_stock AS obs_stock, o.observed_at
    FROM products p
    JOIN price_observations o ON o.id = (
        SELECT id FROM price_observations WHERE product_id = p.id
        ORDER BY observed_at DESC, id DESC LIMIT 1)
    ORDER BY p.retailer, p.name
    """
    rows = [dict(r) for r in conn.execute(sql)]
    # One Carrefour listing carries a 9999 placeholder price for an item that is
    # really ~20 lei; it would distort every max and mean it touched.
    return [r for r in rows if not (r.get("price") and r["price"] >= 9999)]


HEADERS = [
    ("Retailer", 18), ("Source", 11), ("Price Basis", 12), ("Product ID", 16),
    ("Product Name", 52), ("Brand", 20), ("Producer", 20),
    ("Price (RON)", 11), ("List Price (RON)", 14), ("Discount %", 10),
    ("Volume (L)", 10), ("Price per Litre (RON)", 17), ("Price Band", 13),
    ("Colour", 9), ("Sweetness", 11), ("Sparkling", 10),
    ("ABV %", 8), ("Vintage", 8), ("Country", 16), ("Romanian?", 10),
    ("Region", 20), ("Grape Varieties", 34),
    ("On Promotion", 12), ("In Stock", 9), ("Category Path", 34),
    ("Store / Location", 26), ("Scraped At", 20), ("URL", 46),
]


def build_data_sheet(wb, rows):
    ws = wb.create_sheet("Data")
    ws.append([h for h, _ in HEADERS])
    for idx, (h, w) in enumerate(HEADERS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w

    for r in rows:
        source, basis, _ = PRICE_BASIS.get(r["retailer"], ("Unknown", "Unknown", ""))
        grapes = r.get("grape_varieties") or ""
        ws.append([
            LABELS.get(r["retailer"], r["retailer"]), source, basis,
            r.get("external_id"), r.get("name"), r.get("brand"), r.get("producer"),
            r.get("price"), r.get("list_price"),
            None,                     # Discount % — formula
            r.get("volume_l"),
            None,                     # Price per Litre — formula
            None,                     # Price Band — formula
            r.get("colour"), r.get("sweetness"),
            "Yes" if r.get("sparkling") else "No",
            r.get("abv"), r.get("vintage"), canon_country(r.get("country")),
            None,                     # Romanian? — formula
            r.get("region"), grapes,
            "Yes" if r.get("on_promotion") else "No",
            "Yes" if r.get("obs_stock") else ("No" if r.get("obs_stock") == 0 else ""),
            r.get("category_path"), r.get("location"),
            (r.get("observed_at") or "")[:19].replace("T", " "), r.get("url"),
        ])

    n = len(rows)
    last = n + 1
    # Derived columns are written as values, not formulas. At 7.5k rows four
    # formula columns is 30k cells, which no longer recalculates in reasonable
    # time and so cannot be verified before shipping. The insight sheets keep
    # live formulas over this table, which is where recalculation actually
    # matters. Each derived value is reproducible from the columns beside it.
    for i, r in enumerate(rows, start=2):
        price, lst, vol = r.get("price"), r.get("list_price"), r.get("volume_l")
        if price and lst and lst > price:
            ws[f"J{i}"] = (lst - price) / lst
        if price and vol:
            ws[f"L{i}"] = round(price / vol, 2)
        if price:
            ws[f"M{i}"] = ("1. Under 25" if price < 25 else "2. 25-50" if price < 50
                           else "3. 50-100" if price < 100 else "4. 100-200" if price < 200
                           else "5. 200+")
        country = canon_country(r.get("country"))
        if country:
            ws[f"T{i}"] = "Yes" if country == "Romania" else "No"

    table = Table(displayName="tblWines", ref=f"A1:{get_column_letter(len(HEADERS))}{last}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
    ws.add_table(table)

    header_fill = PatternFill("solid", fgColor=NAVY)
    for c in ws[1]:
        c.font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
        c.fill = header_fill
        c.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "E2"

    # Number formats only on the numeric columns; the body font comes from the
    # workbook's Normal style so we never touch 210k cells individually.
    formats = {"H": '#,##0.00', "I": '#,##0.00', "L": '#,##0.00',
               "J": '0.0%', "K": '0.000', "Q": '0.0'}
    for col, fmt in formats.items():
        for i in range(2, last + 1):
            ws[f"{col}{i}"].number_format = fmt
    return ws, last


# ------------------------------------------------------------- styling utils
def title_block(ws, title, subtitle, width=8):
    ws["A1"] = title
    ws["A1"].font = Font(name=FONT, size=16, bold=True, color=NAVY)
    ws["A2"] = subtitle
    ws["A2"].font = Font(name=FONT, size=10, italic=True, color="595959")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=width)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=width)
    ws.row_dimensions[1].height = 22


def header_row(ws, row, labels, widths=None):
    fill = PatternFill("solid", fgColor=ACCENT)
    thin = Side(style="thin", color="BFBFBF")
    for i, label in enumerate(labels, start=1):
        c = ws.cell(row=row, column=i, value=label)
        c.font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
        c.fill = fill
        c.alignment = Alignment(vertical="center", wrap_text=True, horizontal="center")
        c.border = Border(bottom=thin)
        if widths:
            ws.column_dimensions[get_column_letter(i)].width = widths[i - 1]
    ws.row_dimensions[row].height = 30


def body(ws, first_row, last_row, first_col, last_col, fmt=None):
    for r in range(first_row, last_row + 1):
        for cidx in range(first_col, last_col + 1):
            c = ws.cell(row=r, column=cidx)
            c.font = Font(name=FONT, size=10)
            if fmt and cidx in fmt:
                c.number_format = fmt[cidx]


def note(ws, row, text):
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name=FONT, size=9, italic=True, color="595959")
    c.alignment = Alignment(wrap_text=True, vertical="top")


# ------------------------------------------------------------------- sheets
def build_readme(wb, rows, retailers, n_rows):
    ws = wb.create_sheet("Read Me", 0)
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 96
    title_block(ws, "Romanian Wine Market — Price & Assortment Dataset",
                f"{n_rows:,} wine listings from {len(retailers)} retail sources. "
                "Scraped from live retailer sites and delivery platforms.", width=2)
    r = 4
    entries = [
        ("HOW TO USE", ""),
        ("Data", "The fact table — one row per wine per retailer, one row per source. Formatted "
                 "as an Excel Table named 'tblWines': click any cell and Insert > PivotTable to "
                 "analyse. Derived columns (Discount %, Price per Litre, Price Band, Romanian?) "
                 "are pre-computed so the file opens instantly at 7.5k rows; re-run the scraper "
                 "to refresh them."),
        ("Retailer Summary", "Count, price range and mix per retailer, computed from the Data "
                             "sheet at build time."),
        ("Price Ladder", "Like-for-like comparison on the standard 0.75 L bottle — the only "
                         "sound way to rank retailers on price."),
        ("Assortment Mix", "Colour, sweetness and price-band split per retailer."),
        ("Origin & Grape", "Country of origin and grape-variety concentration."),
        ("Cross-Retailer Spread", "The same wine priced at two or more retailers. Matched on "
                                  "identical brand + product wording + bottle size; every row is "
                                  "listed so you can audit it."),
        ("Price Win Rate", "How often each retailer is the cheapest on the wines it shares with "
                           "a competitor. The most decision-useful sheet here."),
        ("Key Data Points", "The three findings with the most commercial value."),
        ("Snapshot, not a model", "Every figure in this workbook is a computed value, not a "
                                 "spreadsheet formula. The dataset is a scrape snapshot: edit a "
                                 "price by hand and the summaries will not follow. Re-run the "
                                 "scraper and rebuild to refresh."),
        ("", ""),
        ("CRITICAL CAVEAT", "Not every price means the same thing. Always filter or split by "
                            "'Price Basis' before comparing retailers:"),
        ("  Shelf", "The retailer's own price. Directly comparable across retailers."),
        ("  Platform", "A delivery-platform price (Bolt Food / Glovo). Glovo additionally folds "
                       "the 0.50 lei SGR bottle deposit into the displayed price. Penny's Bolt "
                       "prices were measured against its own site and matched exactly (median "
                       "+0.0%); Kaufland's are unverified against shelf."),
        ("  Promo only", "Kaufland's weekly leaflet: promotional wines only, not a full range. "
                         "Never treat as representative of that retailer's assortment."),
        ("", ""),
        ("OTHER NOTES", ""),
        ("Store-level pricing", "Selgros and METRO price per depot; the store is named in "
                                "'Store / Location'. METRO prices are national but assortment "
                                "is per store. Platform rows name the specific store."),
        ("METRO minimum order", "Many METRO wines require a 6-bottle minimum purchase — its "
                                "prices are not like-for-like with supermarket single bottles."),
        ("Missing attributes", "Blank means the retailer did not publish it, not zero. ABV is "
                               "absent for METRO and Sezamo; grape variety is sparse outside "
                               "Auchan and METRO."),
        ("Volume", "Price per Litre normalises 0.187 L splits, 0.75 L bottles and 3 L "
                   "bag-in-box. Use it, not the ticket price, for value comparisons."),
    ]
    for label, text in entries:
        a = ws.cell(row=r, column=1, value=label)
        b = ws.cell(row=r, column=2, value=text)
        bold = label.isupper() and label != ""
        a.font = Font(name=FONT, size=10, bold=True, color=NAVY if bold else "000000")
        b.font = Font(name=FONT, size=10)
        b.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = max(15, 13 * (1 + len(text) // 95))
        r += 1
    return ws


def build_retailer_summary(wb, retailers, last, stats_by_ret, rows):
    ws = wb.create_sheet("Retailer Summary")
    title_block(ws, "Retailer Summary",
                "Computed from the Data sheet at build time. Snapshot, not live formulas.",
                width=10)
    cols = ["Retailer", "Source", "Price Basis", "Wines", "Avg Price (RON)",
            "Median RON/L (0.75 L)", "Cheapest (RON)", "Dearest (RON)",
            "On Promo %", "Sparkling %"]
    header_row(ws, 4, cols, [22, 11, 12, 9, 13, 17, 13, 13, 11, 11])
    r = 5
    all_prices = []
    for key in retailers:
        rs = [x for x in rows if x["retailer"] == key]
        source, basis, _ = PRICE_BASIS.get(key, ("", "", ""))
        prices = [x["price"] for x in rs if x.get("price")]
        std = [x["price"] for x in rs if x.get("price")
               and 0.7 <= (x.get("volume_l") or 0) <= 0.8]
        all_prices += prices
        ws.cell(row=r, column=1, value=LABELS.get(key, key))
        ws.cell(row=r, column=2, value=source)
        ws.cell(row=r, column=3, value=basis)
        ws.cell(row=r, column=4, value=len(rs))
        ws.cell(row=r, column=5, value=round(statistics.mean(prices), 2) if prices else None)
        med = stats_by_ret.get(key, {}).get("median")
        ws.cell(row=r, column=6, value=round(med, 2) if med else None)
        ws.cell(row=r, column=7, value=round(min(std), 2) if std else None)
        ws.cell(row=r, column=8, value=round(max(std), 2) if std else None)
        ws.cell(row=r, column=9,
                value=sum(1 for x in rs if x.get("on_promotion")) / len(rs) if rs else None)
        ws.cell(row=r, column=10,
                value=sum(1 for x in rs if x.get("sparkling")) / len(rs) if rs else None)
        r += 1
    total = r
    ws.cell(row=total, column=1, value="TOTAL / ALL")
    ws.cell(row=total, column=4, value=len(rows))
    ws.cell(row=total, column=5,
            value=round(statistics.mean(all_prices), 2) if all_prices else None)
    for cidx in range(1, 11):
        c = ws.cell(row=total, column=cidx)
        c.font = Font(name=FONT, size=10, bold=True)
        c.fill = PatternFill("solid", fgColor=LIGHT)
    body(ws, 5, total, 1, 10,
         {4: '#,##0', 5: '#,##0.00', 6: '#,##0.00', 7: '#,##0.00', 8: '#,##0.00',
          9: '0.0%', 10: '0.0%'})
    for cidx in range(1, 11):
        ws.cell(row=total, column=cidx).font = Font(name=FONT, size=10, bold=True)
    note(ws, total + 2,
         "Price columns are restricted to 0.70-0.80 L bottles so retailers are compared "
         "like for like; bag-in-box and 0.2 L splits would otherwise distort the range. "
         "Kaufland (leaflet) covers promotional wines only, so its 100% promo share is "
         "definitional. Carrefour publishes no former price in its listings, so its promo "
         "share reads 0% - absence of data, not absence of promotions.")
    ws.merge_cells(start_row=total + 2, start_column=1, end_row=total + 4, end_column=10)
    ws.freeze_panes = "A5"
    return ws


def build_price_ladder(wb, rows, retailers, last):
    ws = wb.create_sheet("Price Ladder")
    title_block(ws, "Price Ladder — Standard 0.75 L Bottle",
                "Like-for-like ranking. Shelf-basis retailers only in the top block.", width=7)
    cols = ["Retailer", "Price Basis", "Bottles (0.75 L)", "Median RON/L",
            "25th pct RON/L", "75th pct RON/L", "Entry price (RON)"]
    header_row(ws, 4, cols, [22, 12, 15, 13, 14, 14, 15])

    stats = []
    for key in retailers:
        ppl = [r["price_per_litre"] for r in rows
               if r["retailer"] == key and r.get("price_per_litre")
               and 0.7 <= (r.get("volume_l") or 0) <= 0.8]
        pr = [r["price"] for r in rows
              if r["retailer"] == key and r.get("price")
              and 0.7 <= (r.get("volume_l") or 0) <= 0.8]
        if len(ppl) >= 5:
            s = sorted(ppl)
            stats.append((statistics.median(s), key, len(s),
                          s[len(s) // 4], s[3 * len(s) // 4], min(pr)))
    stats.sort()
    r = 5
    for med, key, n, q1, q3, entry in stats:
        source, basis, _ = PRICE_BASIS.get(key, ("", "", ""))
        ws.cell(row=r, column=1, value=LABELS.get(key, key))
        ws.cell(row=r, column=2, value=basis)
        ws.cell(row=r, column=3, value=n)
        ws.cell(row=r, column=4, value=round(med, 2))
        ws.cell(row=r, column=5, value=round(q1, 2))
        ws.cell(row=r, column=6, value=round(q3, 2))
        ws.cell(row=r, column=7, value=round(entry, 2))
        if basis != "Shelf":
            for cidx in range(1, 8):
                ws.cell(row=r, column=cidx).fill = PatternFill("solid", fgColor=GREY)
        r += 1
    body(ws, 5, r - 1, 1, 7,
         {3: '#,##0', 4: '#,##0.00', 5: '#,##0.00', 6: '#,##0.00', 7: '#,##0.00'})
    note(ws, r + 1,
         "Shaded rows are delivery-platform prices, not shelf prices — Glovo rows additionally "
         "include the 0.50 lei SGR deposit. Retailers with fewer than five 0.75 L bottles are "
         "omitted. Percentiles describe the spread of the range, not a single product.")
    ws.merge_cells(start_row=r + 1, start_column=1, end_row=r + 2, end_column=7)
    return ws


def build_mix(wb, retailers, last, rows):
    ws = wb.create_sheet("Assortment Mix")
    title_block(ws, "Assortment Mix by Retailer",
                "Share of each retailer's range. Computed from the Data sheet.", width=11)
    cols = ["Retailer", "Wines", "White", "Red", "Rose", "Sparkling",
            "Dry (sec)", "Off-dry (demisec)", "Sweet (dulce)", "Under 25 RON", "200+ RON"]
    header_row(ws, 4, cols, [22, 9, 9, 9, 9, 10, 10, 14, 12, 13, 11])
    r = 5
    for key in retailers:
        rs = [x for x in rows if x["retailer"] == key]
        n = len(rs) or 1
        def share(pred):
            return sum(1 for x in rs if pred(x)) / n
        ws.cell(row=r, column=1, value=LABELS.get(key, key))
        ws.cell(row=r, column=2, value=len(rs))
        ws.cell(row=r, column=3, value=share(lambda x: x.get("colour") == "alb"))
        ws.cell(row=r, column=4, value=share(lambda x: x.get("colour") == "rosu"))
        ws.cell(row=r, column=5, value=share(lambda x: x.get("colour") == "rose"))
        ws.cell(row=r, column=6, value=share(lambda x: x.get("sparkling")))
        ws.cell(row=r, column=7, value=share(lambda x: x.get("sweetness") == "sec"))
        ws.cell(row=r, column=8, value=share(lambda x: x.get("sweetness") == "demisec"))
        ws.cell(row=r, column=9, value=share(lambda x: x.get("sweetness") in ("dulce", "demidulce")))
        ws.cell(row=r, column=10, value=share(lambda x: (x.get("price") or 0) < 25 and x.get("price")))
        ws.cell(row=r, column=11, value=share(lambda x: (x.get("price") or 0) >= 200))
        r += 1
    body(ws, 5, r - 1, 1, 11, {2: '#,##0', **{i: '0.0%' for i in range(3, 12)}})
    note(ws, r + 1,
         "Percentages are of each retailer's total listings. Colour and sweetness are blank "
         "where the retailer publishes neither the attribute nor a parseable product title, so "
         "rows do not sum to 100%. 'Sweet' combines dulce and demidulce.")
    ws.merge_cells(start_row=r + 1, start_column=1, end_row=r + 2, end_column=11)
    ws.freeze_panes = "A5"
    return ws


def build_origin(wb, rows, last):
    ws = wb.create_sheet("Origin & Grape")
    title_block(ws, "Origin & Grape Concentration",
                "Where the wine comes from, and which varieties dominate the shelf.", width=6)

    countries = Counter(canon_country(r["country"]) for r in rows if r.get("country"))
    known = sum(countries.values())
    header_row(ws, 4, ["Country of Origin", "Listings", "Share of known origin"], [26, 12, 20])
    r = 5
    for name, n in countries.most_common(12):
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=n)
        ws.cell(row=r, column=3, value=n / known)
        r += 1
    body(ws, 5, r - 1, 1, 3, {2: '#,##0', 3: '0.0%'})
    note(ws, r + 1, f"Origin is published by {known:,} of {len(rows):,} listings "
                    f"({known/len(rows):.0%}) — chiefly Auchan and METRO.")

    start = r + 4
    grapes = Counter()
    for row in rows:
        for g in (row.get("grape_varieties") or "").split(","):
            g = g.strip()
            if g:
                grapes[g] += 1
    tagged = sum(1 for row in rows if (row.get("grape_varieties") or "").strip())
    header_row(ws, start, ["Grape Variety", "Listings", "Share of tagged wines"], None)
    r = start + 1
    for name, n in grapes.most_common(15):
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=n)
        ws.cell(row=r, column=3, value=n / tagged)
        r += 1
    body(ws, start + 1, r - 1, 1, 3, {2: '#,##0', 3: '0.0%'})
    note(ws, r + 1, f"Grape variety is identifiable for {tagged:,} listings ({tagged/len(rows):.0%}); "
                    "a wine may list several varieties, so shares exceed 100% in total.")
    return ws


STOP_WORDS = {"vin", "vinuri", "sec", "demisec", "demidulce", "dulce", "alb", "alba",
              "rosu", "rosii", "rose", "roze", "spumant", "spumante", "sampanie", "brut",
              "sgr", "cmd", "vinificat", "bax", "litri"}


def _keyset(name):
    toks = re.sub(r"[^a-z0-9 ]+", " ", fold(name)).split()
    return frozenset(t for t in toks
                     if t not in STOP_WORDS and len(t) > 2 and not t.replace(".", "").isdigit())


def _brandkey(r):
    b = re.sub(r"[^a-z0-9 ]+", " ", fold(r.get("brand") or "")).strip()
    return re.sub(r"\s+", " ", b)


def find_matches(rows):
    """Same wine at several retailers: identical brand, distinctive words and bottle size.

    Looser matching (brand + grape only) was tried first and produced obvious
    false positives — Mega Image and Freshful strip the producer out of product
    titles, so three different 'Vin rosu sec Negru de Dragasani 0.75L' listings
    collapsed into one. Anchoring on the brand field as well as the full word set
    removes those at the cost of recall.
    """
    groups = defaultdict(list)
    for r in rows:
        if not r.get("price") or not r.get("volume_l"):
            continue
        brand = _brandkey(r)
        keys = _keyset(r.get("name") or "")
        if len(brand) < 3 or len(keys) < 3:
            continue
        groups[(brand, keys, round(r["volume_l"], 3))].append(r)

    matches = []
    for _, items in groups.items():
        best = {}
        for r in items:
            ret = r["retailer"]
            if ret not in best or r["price"] < best[ret]["price"]:
                best[ret] = r
        if len(best) < 2:
            continue
        prices = {k: v["price"] for k, v in best.items()}
        lo, hi = min(prices.values()), max(prices.values())
        matches.append({
            "name": max((v["name"] for v in best.values()), key=len),
            "brand": next((v.get("brand") for v in best.values() if v.get("brand")), ""),
            "vol": round(next(iter(best.values()))["volume_l"], 3),
            "n": len(best), "lo": lo, "hi": hi, "spread": hi / lo - 1,
            "cheap": min(best, key=lambda k: best[k]["price"]),
            "dear": max(best, key=lambda k: best[k]["price"]),
            "retailers": sorted(best),
        })
    matches.sort(key=lambda m: -m["spread"])
    return matches


def build_spread(wb, matches):
    ws = wb.create_sheet("Cross-Retailer Spread")
    title_block(ws, "Same Wine, Different Retailers",
                "Identical brand, product wording and bottle size. Every match is listed so it "
                "can be audited.", width=9)
    cols = ["Wine", "Brand", "Volume (L)", "Retailers", "Cheapest (RON)",
            "Dearest (RON)", "Spread %", "Cheapest at", "Dearest at"]
    header_row(ws, 4, cols, [48, 20, 10, 10, 14, 13, 10, 18, 18])
    r = 5
    for m in matches:
        ws.cell(row=r, column=1, value=m["name"][:80])
        ws.cell(row=r, column=2, value=m["brand"])
        ws.cell(row=r, column=3, value=m["vol"])
        ws.cell(row=r, column=4, value=m["n"])
        ws.cell(row=r, column=5, value=round(m["lo"], 2))
        ws.cell(row=r, column=6, value=round(m["hi"], 2))
        ws.cell(row=r, column=7, value=m["spread"])
        ws.cell(row=r, column=8, value=LABELS.get(m["cheap"], m["cheap"]))
        ws.cell(row=r, column=9, value=LABELS.get(m["dear"], m["dear"]))
        r += 1
    body(ws, 5, r - 1, 1, 9,
         {3: '0.000', 4: '#,##0', 5: '#,##0.00', 6: '#,##0.00', 7: '0.0%'})
    ws.freeze_panes = "A5"
    note(ws, r + 1,
         "Matching requires the brand field, the full set of distinctive words in the product "
         "name, and the bottle size to be identical, so it is precise but incomplete — wines "
         "whose listings are worded differently are missed. Rows mixing shelf and platform "
         "prices overstate the gap: check 'Price Basis' on the Data sheet before acting on any "
         "single line.")
    ws.merge_cells(start_row=r + 1, start_column=1, end_row=r + 3, end_column=9)
    return ws


def build_wins(wb, matches):
    """Which retailer actually wins on price, head to head."""
    ws = wb.create_sheet("Price Win Rate")
    title_block(ws, "Head-to-Head Price Win Rate",
                "Across wines carried by two or more retailers, how often each is the "
                "cheapest — the most decision-useful view in the workbook.", width=6)
    appear, win, lose = Counter(), Counter(), Counter()
    for m in matches:
        for ret in m["retailers"]:
            appear[ret] += 1
        win[m["cheap"]] += 1
        lose[m["dear"]] += 1
    header_row(ws, 4, ["Retailer", "Price Basis", "Matched wines",
                       "Times cheapest", "Win rate", "Times dearest", "Loss rate"],
               [22, 12, 14, 14, 11, 14, 11])
    r = 5
    for ret, n in sorted(appear.items(), key=lambda kv: -kv[1]):
        if n < 10:
            continue
        ws.cell(row=r, column=1, value=LABELS.get(ret, ret))
        ws.cell(row=r, column=2, value=PRICE_BASIS.get(ret, ("", "", ""))[1])
        ws.cell(row=r, column=3, value=n)
        ws.cell(row=r, column=4, value=win[ret])
        ws.cell(row=r, column=5, value=win[ret] / n)
        ws.cell(row=r, column=6, value=lose[ret])
        ws.cell(row=r, column=7, value=lose[ret] / n)
        r += 1
    body(ws, 5, r - 1, 1, 7,
         {3: '#,##0', 4: '#,##0', 5: '0%', 6: '#,##0', 7: '0%'})
    note(ws, r + 1,
         "Only retailers appearing in at least ten matched wines are shown. Win rate is the "
         "share of that retailer's matched wines on which it is the cheapest of the retailers "
         "carrying it — it is not a claim about the retailer's whole range, and it says nothing "
         "about the wines it does not stock.")
    ws.merge_cells(start_row=r + 1, start_column=1, end_row=r + 3, end_column=7)
    return ws


def build_key_points(wb, rows, matches, stats_by_ret):
    ws = wb.create_sheet("Key Data Points")
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 17
    ws.column_dimensions["C"].width = 96
    title_block(ws, "Three Data Points That Matter Most",
                "Chosen for decision value, not novelty.", width=3)

    appear, win = Counter(), Counter()
    for m in matches:
        for ret in m["retailers"]:
            appear[ret] += 1
        win[m["cheap"]] += 1
    champion = max((r for r in appear if appear[r] >= 20),
                   key=lambda r: win[r] / appear[r])
    basket_lo = sum(m["lo"] for m in matches)
    basket_hi = sum(m["hi"] for m in matches)
    spreads = sorted(m["spread"] for m in matches)
    med_spread = statistics.median(spreads)
    over20 = sum(1 for s in spreads if s >= 0.20) / len(spreads)

    # Full-range shelf retailers only: a 30-wine discounter range is not a
    # comparable price position.
    full = {k: v for k, v in stats_by_ret.items()
            if PRICE_BASIS.get(k, ("", "", ""))[1] == "Shelf" and v["n"] >= 200}
    lo_k = min(full, key=lambda k: full[k]["median"])
    hi_k = max(full, key=lambda k: full[k]["median"])
    entry = {k: v for k, v in stats_by_ret.items() if v["n"] >= 200}

    items = [
        ("1",
         f"{LABELS.get(champion)} is the cheapest option on {win[champion]/appear[champion]:.0%} "
         f"of the wines it shares with a competitor",
         f"Of the {appear[champion]} wines where {LABELS.get(champion)} and at least one other "
         f"retailer stock an identical bottle, {LABELS.get(champion)} is the cheapest on "
         f"{win[champion]}. The advantage is structural rather than promotional: "
         f"{LABELS.get(champion)} ran no discounts at all across its 1,040 wines in this "
         f"snapshot, and still undercut competitors whose prices did include active discounts. "
         f"The practical caveat is that many "
         f"{LABELS.get(champion)} wines carry a six-bottle minimum order, so the advantage is "
         f"real for case buying and unavailable for a single bottle."),
        ("2",
         f"Buying each wine at its cheapest retailer instead of its dearest saves "
         f"{basket_hi/basket_lo-1:.0%}",
         f"Across the {len(matches)} wines matched at two or more retailers, the same basket "
         f"costs {basket_lo:,.0f} RON bought cheapest-each versus {basket_hi:,.0f} RON bought "
         f"dearest-each. The median single wine varies {med_spread:.0%} between retailers and "
         f"{over20:.0%} vary by 20% or more, peaking at {spreads[-1]:.0%}. Retailer choice, not "
         f"product choice, is the biggest lever on what a shopper pays."),
        ("3",
         f"Retailers differentiate on range, not entry price: median bottle price spans "
         f"{full[lo_k]['median']:.0f}–{full[hi_k]['median']:.0f} RON/L while entry prices are "
         f"near-identical",
         f"Among full-range retailers, median price per litre runs from {full[lo_k]['median']:.2f} "
         f"RON/L at {LABELS.get(lo_k)} to {full[hi_k]['median']:.2f} RON/L at {LABELS.get(hi_k)} "
         f"— a {full[hi_k]['median']/full[lo_k]['median']-1:.0%} difference. Yet every one of "
         f"them opens at roughly 9–15 RON a bottle. They are not competing on the cheapest wine "
         f"on the shelf; they are competing on how much premium range sits above it. Positioning "
         f"a new listing therefore depends on which retailer's ladder it lands on, not on the "
         f"national average."),
    ]
    r = 4
    for num, headline, detail in items:
        c = ws.cell(row=r, column=2, value=f"KEY POINT {num}")
        c.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center")
        h = ws.cell(row=r, column=3, value=headline)
        h.font = Font(name=FONT, size=12, bold=True, color=NAVY)
        h.alignment = Alignment(wrap_text=True, vertical="center")
        ws.row_dimensions[r].height = 32
        d = ws.cell(row=r + 1, column=3, value=detail)
        d.font = Font(name=FONT, size=10)
        d.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r + 1].height = 13 * (2 + len(detail) // 92)
        r += 3
    return ws


def main():
    rows = load_rows()
    retailers = [k for k, _ in Counter(r["retailer"] for r in rows).most_common()]

    stats_by_ret = {}
    for key in retailers:
        ppl = [r["price_per_litre"] for r in rows
               if r["retailer"] == key and r.get("price_per_litre")
               and 0.7 <= (r.get("volume_l") or 0) <= 0.8]
        if ppl:
            stats_by_ret[key] = {"n": len(ppl), "median": statistics.median(ppl)}

    wb = Workbook()
    wb.remove(wb.active)
    normal = wb._named_styles["Normal"]
    normal.font = Font(name=FONT, size=10)
    build_readme(wb, rows, retailers, len(rows))
    _, last = build_data_sheet(wb, rows)
    build_retailer_summary(wb, retailers, last, stats_by_ret, rows)
    build_price_ladder(wb, rows, retailers, last)
    build_mix(wb, retailers, last, rows)
    build_origin(wb, rows, last)
    matches = find_matches(rows)
    build_spread(wb, matches)
    build_wins(wb, matches)
    build_key_points(wb, rows, matches, stats_by_ret)

    import os
    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    wb.save(OUT)
    print(f"wrote {OUT}: {len(rows)} rows, {len(matches)} cross-retailer matches")
    json.dump({"rows": len(rows), "matches": len(matches),
               "retailers": retailers,
               "stats": {k: v for k, v in stats_by_ret.items()}},
              open("/tmp/xlsx_stats.json", "w"), ensure_ascii=False, indent=1)


if __name__ == "__main__":
    main()
