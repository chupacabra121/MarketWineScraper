"""Entry-segment price matrix: 2 litre brands under 10 lei a litre, store by store.

One table. Rows are stores, columns are brands, and a cell is the cheapest that
brand gets in that store — pack price with the litre price beside it.

Only the 2 litre PET is compared. The segment also holds 3 litre boxes, 5 and 10
litre casks, and one 1 litre; letting them in put a 79.98-lei cask beside a
15.40-lei bottle in the same column, which reads as a price gap and is a format
gap. Setting ``FORMATS`` to another size, or to None, re-cuts the same table.

The threshold reads the list price *including* the SGR deposit, which is what a
shopper hands over. At 2 litres the deposit is 0.25 lei a litre, enough to move
a wine across a 10-lei line on its own.
"""
from __future__ import annotations

import os
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from winescraper import identity, pricing
from winescraper.storage import Store

DB = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                  "data", "wines.sqlite")
OUT = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                   "exports", "entry-segment-under-10.xlsx")

CEILING = 10.0
#: Bottle sizes to compare. ``None`` lets every format in.
FORMATS: tuple[float, ...] | None = (2.0,)

STORES = {
    "metro": "METRO", "selgros": "Selgros", "carrefour": "Carrefour",
    "penny_bolt": "Penny (Bolt)", "kaufland_bolt": "Kaufland (Bolt)",
    "profi_glovo": "Profi (Glovo)", "auchan": "Auchan", "penny": "Penny",
    "freshful": "Freshful", "mega_image": "Mega Image", "sezamo": "Sezamo",
    "supeco_glovo": "Supeco (Glovo)", "kaufland": "Kaufland (pliant)",
}

# Brands each shop writes its own way. Merging them is an assertion — the data
# only shows the words on the shelf — so every one is listed explicitly rather
# than inferred from a similarity rule that would also merge things it should
# not. Keyed by a word that has to appear in the folded title.
MERGED = {
    "babanu": "Babanu",                  # Selgros, METRO "Murfatlar Babanu", Profi "Babanul"
    "muscatel": "Muscatel",              # Selgros, METRO "Beciul Domnesc Muscatel"
    "stramosesc": "Vinul Strămoșesc",    # Selgros, METRO "Rovinex Vin Stramosesc"
    "poloboace": "9 Poloboace",          # Rovinex's other entry range, kept apart
    "carisma": "El Carisma",             # one shop, split by the 1 L / 3 L wording
    "ulcior": "Ulcior",                  # Penny's own table wine, "Alb/Rosu de ulcior"
}

# Display names for the rest, where the automatic reading is right but terse.
RENAMED = {
    "perla hangitei": "Perla Hangiței",
    "drumul vie": "Drumul de la Vie",
    "beciul podgoreanului": "Beciul Podgoreanului",
    "dealurile vinului": "Dealurile Vinului",
    "val duna": "Val Duna",
    "amigo": "Vino Amigo",
    "premiat": "Premiat (Vincon)",
    "vinexport": "Premiat (Vinexport)",
}

INK, RULE, HEAD = "231F20", "757575", "F2F2F2"


def brand_of(row, lexicon) -> str:
    folded = identity.expand(f"{row['name']} {row.get('brand') or ''}")
    for token, label in MERGED.items():
        if token in folded:
            return label
    sig = identity.signature(row, lexicon)
    label = sig.brand or " ".join(sorted(sig.anchor)) or "(fără marcă)"
    return RENAMED.get(label, label.title())


def main() -> None:
    store = Store(DB)
    rows = [dict(r) for r in store.latest()]
    store.close()

    lexicon = identity.brand_lexicon(rows)
    segment = []
    for row in rows:
        volume = row.get("volume_l")
        if FORMATS and (volume is None
                        or not any(abs(volume - f) < 0.01 for f in FORMATS)):
            continue
        price = pricing.regular(row)
        ppl = pricing.per_litre(price, volume)
        if ppl is None or ppl >= CEILING:
            continue
        segment.append((brand_of(row, lexicon), row["retailer"], price, ppl))

    # Cheapest listing of each brand in each store.
    best: dict[tuple[str, str], tuple[float, float]] = {}
    for brand, retailer, price, ppl in segment:
        key = (brand, retailer)
        if key not in best or ppl < best[key][1]:
            best[key] = (price, ppl)

    brands = sorted({b for b, _ in best},
                    key=lambda b: (-len({r for (x, r) in best if x == b}),
                                   min(v[1] for (x, _), v in best.items() if x == b)))
    retailers = sorted({r for _, r in best},
                       key=lambda r: (-len({b for (b, x) in best if x == r}),
                                      STORES.get(r, r)))

    book = Workbook()
    sheet = book.active
    sheet.title = "Sub 10 lei pe litru"
    numeric = book.create_sheet("Lei pe litru")

    for ws, combined in ((sheet, True), (numeric, False)):
        sizes = ("toate formatele" if not FORMATS
                 else " / ".join(f"{f:g} L" for f in FORMATS))
        ws["A1"] = (f"Vinuri de {sizes} sub 10 lei/litru — preț de raft, SGR inclus"
                    if combined else "Aceleași date, doar lei/litru (numeric)")
        ws["A1"].font = Font(name="Georgia", size=13, bold=True, color=INK)
        ws["A2"] = ("Celula: prețul pachetului, iar în paranteză prețul pe litru. "
                    "Cel mai ieftin sortiment al mărcii în magazinul respectiv."
                    if combined else
                    "Cel mai ieftin sortiment al mărcii în magazinul respectiv.")
        ws["A2"].font = Font(name="Arial", size=9, italic=True, color=RULE)

        head = 4
        ws.cell(head, 1, "Magazin").font = Font(name="Arial", size=9, bold=True, color=INK)
        for n, brand in enumerate(brands, start=2):
            cell = ws.cell(head, n, brand)
            cell.font = Font(name="Arial", size=9, bold=True, color=INK)
            cell.alignment = Alignment(horizontal="center", wrap_text=True, vertical="bottom")
            cell.fill = PatternFill("solid", fgColor=HEAD)
        ws.cell(head, 1).fill = PatternFill("solid", fgColor=HEAD)

        for m, retailer in enumerate(retailers, start=head + 1):
            label = ws.cell(m, 1, STORES.get(retailer, retailer))
            label.font = Font(name="Arial", size=9, bold=True, color=INK)
            for n, brand in enumerate(brands, start=2):
                found = best.get((brand, retailer))
                cell = ws.cell(m, n)
                if found:
                    price, ppl = found
                    cell.value = f"{price:.2f} ({ppl:.2f}/L)" if combined else ppl
                    if not combined:
                        cell.number_format = "0.00"
                else:
                    cell.value = "—"
                cell.font = Font(name="Arial", size=9, color=INK)
                cell.alignment = Alignment(horizontal="center")

        edge = Side(style="thin", color=RULE)
        last = head + len(retailers)
        for m in range(head, last + 1):
            for n in range(1, len(brands) + 2):
                ws.cell(m, n).border = Border(top=edge if m == head else None,
                                              bottom=edge if m in (head, last) else None)
        ws.column_dimensions["A"].width = 18
        for n in range(2, len(brands) + 2):
            ws.column_dimensions[get_column_letter(n)].width = 16 if combined else 11
        ws.freeze_panes = ws.cell(head + 1, 2)

    book.save(OUT)
    print(f"wrote {OUT}: {len(retailers)} magazine x {len(brands)} mărci, "
          f"{len(segment)} listări sub {CEILING:.0f} lei/L"
          + (f" la {'/'.join(f'{f:g} L' for f in FORMATS)}" if FORMATS else ""))


if __name__ == "__main__":
    main()
