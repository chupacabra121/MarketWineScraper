"""Price matrix for a named list of entry brands, split PET against bag-in-box.

Two tables of the same shape: stores down the side, brands across the top, the
cheapest listing of that brand in that store in the cell. The brand list is
given rather than discovered — it is someone's competitive set, so a brand we
carry no listing for keeps its column and shows empty. An empty column is the
answer to "where is it sold", not an omission.

PET and bag-in-box are separated because they are different products at
different prices: Vinul Stramosesc is 8.24 lei a litre in a 2 litre PET at
Selgros and 8.00 in the 10 litre box, and averaging those tells you neither.
"""
from __future__ import annotations

import os
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from winescraper import deposit, pricing
from winescraper.normalize import fold
from winescraper.storage import Store

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DB = os.path.join(ROOT, "data", "wines.sqlite")
OUT = os.path.join(ROOT, "exports", "brand-matrix-pet-bib.xlsx")

STORES = {
    "metro": "METRO", "selgros": "Selgros", "carrefour": "Carrefour",
    "penny_bolt": "Penny (Bolt)", "kaufland_bolt": "Kaufland (Bolt)",
    "profi_glovo": "Profi (Glovo)", "auchan": "Auchan", "penny": "Penny",
    "freshful": "Freshful", "mega_image": "Mega Image", "sezamo": "Sezamo",
    "supeco_glovo": "Supeco (Glovo)", "kaufland": "Kaufland (pliant)",
    "froo": "Froo (raft)",
    # No catalogue of our own; present only through the shelf audit.
    "lidl": "Lidl",
}

# (producer, brand, word that must appear in the folded title). The word is
# what matches, so it has to be specific: "rosado" alone pulls in Freixenet
# Cordon Rosado, a Spanish sparkling wine that is not this brand at all.
PET_BRANDS = [
    ("Murfatlar", "Babanu", "babanu"),
    ("Private label", "Winegarden", "winegarden"),
    ("Private label", "Vino Amigo", "vino amigo"),
    ("Beciul Domnesc", "Muscatel", "muscatel"),
    ("Carpatvin", "de Ulcior", "ulcior"),
    ("Rovinex", "Vinul Strămoșesc", "stramosesc"),
    ("Rovinex", "9 Poloboace", "poloboace"),
    ("Vinexport", "Perla Hangiței", "hangitei"),
    ("Vinexport", "Weisser", "weisser"),
    ("Caraprodvin", "Aurul Moldovei", "aurul moldovei"),
    ("Vinia", "Beciul Sihastrului", "sihastrului"),
    ("Domeniul Burcilor", "Traianel", "traianel|traienel"),
]

BIB_BRANDS = [
    ("Murfatlar", "Ferma Nouă", "ferma noua"),
    ("Rovinex", "Vinul Strămoșesc", "stramosesc"),
    ("Carpatvin", "Beciul Podgoreanului", "podgoreanului"),
    ("Private label", "Vino Rosado", "vino rosado"),
    ("Private label", "El Carisma", "carisma"),
    ("Private label", "Cape Buffalo", "buffalo"),
    ("Private label", "Monaster", "monaster"),
    ("Private label", "Vita Veche", "vita veche"),
]

#: A PET is the 1-2 litre plastic bottle. Anything three litres or over, or
#: named as a box, is bag-in-box — the same rule the deposit uses, since the
#: state draws the line in the same place.
PET_RANGE = (1.0, 2.0)
BIB_MIN_L = 3.0

# Brands seen on a shelf that our scrape does not reach. Read off the SP-IKA
# audit of August 2026, which visited 11 stores in 9 chains: the prices are
# somebody else's measurement, not ours, so they are marked as such in the
# sheet rather than mixed into the scraped figures. The audit quotes shelf
# prices without the deposit, the same basis our sources publish, so the same
# 0.50 lei is added here to keep the column comparable.
# (brand, store key) -> (price excluding deposit, litres, where)
FROM_AUDIT = {
    ("Babanu", "auchan"): (12.69, 2.0, "Ploiesti; 18.19 la Nord"),
    ("Monaster", "lidl"): (26.99, 1.5, "Bucuresti; auditul da 1,5 L, nu bag-in-box"),
}

# Brands known to be stocked somewhere we have neither a price nor a shelf
# reading. Nothing derives these — each one is here because a person said so,
# and the sheet labels them that way. Add a line to extend.
KNOWN_STOCKED: dict[str, list[str]] = {
    "Babanu": ["carrefour"],
}

INK, RULE, HEAD, GHOST = "231F20", "757575", "F2F2F2", "9A9A9A"
AUDIT_INK, TOLD_INK = "1F4E79", "7F6000"


def is_bib(row) -> bool:
    volume = row.get("volume_l")
    if deposit._BAG_IN_BOX.search(row["name"] or ""):
        return True
    return volume is not None and volume >= BIB_MIN_L


def is_pet(row) -> bool:
    volume = row.get("volume_l")
    return (not is_bib(row) and volume is not None
            and PET_RANGE[0] <= volume <= PET_RANGE[1])


def matrix(rows, brands, keep):
    """``{(brand, retailer): (price, per_litre, volume)}``, cheapest per litre."""
    import re

    best: dict[tuple[str, str], tuple] = {}
    for _producer, brand, pattern in brands:
        for row in rows:
            if not keep(row) or not re.search(pattern, fold(row["name"])):
                continue
            price = pricing.regular(row)
            ppl = pricing.per_litre(price, row.get("volume_l"))
            if ppl is None:
                continue
            key = (brand, row["retailer"])
            if key not in best or ppl < best[key][1]:
                best[key] = (price, ppl, row["volume_l"])
    return best


def write(book, title, note, brands, best, combined):
    ws = book.create_sheet(title)
    ws["A1"] = note
    ws["A1"].font = Font(name="Georgia", size=13, bold=True, color=INK)
    ws["A2"] = ("Celula: prețul pachetului, iar în paranteză prețul pe litru. "
                "Cel mai ieftin sortiment al mărcii în magazinul respectiv. "
                "SGR inclus. Negru = preț cules de noi. "
                "Albastru = de pe raft, din auditul din august 2026, unde noi nu ajungem. "
                "Galben = știm că se vinde, dar nu avem prețul. "
                "Coloană gri = marca nu apare nicăieri."
                if combined else
                "Lei pe litru, SGR inclus. Cel mai ieftin sortiment al mărcii.")
    ws["A2"].font = Font(name="Arial", size=9, italic=True, color=RULE)

    named = {b for _p, b, _x in brands}
    extra = {r for (b, r) in FROM_AUDIT if b in named}
    extra |= {r for b, shops in KNOWN_STOCKED.items() if b in named for r in shops}
    retailers = sorted({r for _, r in best} | extra,
                       key=lambda r: (-len({b for (b, x) in best if x == r}),
                                      STORES.get(r, r)))
    stocked = ({b for b, _ in best} | {b for (b, _r) in FROM_AUDIT}
               | set(KNOWN_STOCKED)) & named

    top, head = 4, 5
    ws.cell(top, 1, "Producător").font = Font(name="Arial", size=8, italic=True, color=RULE)
    ws.cell(head, 1, "Magazin").font = Font(name="Arial", size=9, bold=True, color=INK)
    for n, (producer, brand, _pattern) in enumerate(brands, start=2):
        absent = brand not in stocked
        cell = ws.cell(top, n, producer)
        cell.font = Font(name="Arial", size=8, italic=True, color=RULE)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell = ws.cell(head, n, brand)
        cell.font = Font(name="Arial", size=9, bold=True,
                         color=GHOST if absent else INK)
        cell.alignment = Alignment(horizontal="center", wrap_text=True, vertical="bottom")
        cell.fill = PatternFill("solid", fgColor=HEAD)
    ws.cell(head, 1).fill = PatternFill("solid", fgColor=HEAD)

    for m, retailer in enumerate(retailers, start=head + 1):
        ws.cell(m, 1, STORES.get(retailer, retailer)).font = Font(
            name="Arial", size=9, bold=True, color=INK)
        for n, (_producer, brand, _pattern) in enumerate(brands, start=2):
            cell = ws.cell(m, n)
            found = best.get((brand, retailer))
            audit = FROM_AUDIT.get((brand, retailer))
            told = retailer in KNOWN_STOCKED.get(brand, ())
            if found:
                price, ppl, _volume = found
                cell.value = f"{price:.2f} ({ppl:.2f}/L)" if combined else ppl
                colour = INK
            elif audit:
                shelf, litres, _where = audit
                price = shelf + deposit.AMOUNT
                ppl = round(price / litres, 2)
                cell.value = f"{price:.2f} ({ppl:.2f}/L) raft" if combined else ppl
                colour = AUDIT_INK
            elif told:
                cell.value = "se vinde, fără preț" if combined else None
                colour = TOLD_INK
            else:
                cell.value = "—" if combined else None
                colour = GHOST
            if not combined and isinstance(cell.value, float):
                cell.number_format = "0.00"
            cell.font = Font(name="Arial", size=9, color=colour,
                             italic=bool(not found and (audit or told)))
            cell.alignment = Alignment(horizontal="center")

    edge = Side(style="thin", color=RULE)
    last = head + len(retailers)
    for m in range(head, last + 1):
        for n in range(1, len(brands) + 2):
            ws.cell(m, n).border = Border(top=edge if m == head else None,
                                          bottom=edge if m in (head, last) else None)
    ws.column_dimensions["A"].width = 18
    for n in range(2, len(brands) + 2):
        ws.column_dimensions[get_column_letter(n)].width = 16 if combined else 12
    ws.freeze_panes = ws.cell(head + 1, 2)
    return len(retailers), len(stocked)


def main() -> None:
    store = Store(DB)
    rows = [dict(r) for r in store.latest()]
    store.close()

    pet = matrix(rows, PET_BRANDS, is_pet)
    bib = matrix(rows, BIB_BRANDS, is_bib)

    book = Workbook()
    book.remove(book.active)
    shops_pet, have_pet = write(book, "PET", "Mărci la PET — preț de raft, SGR inclus",
                                PET_BRANDS, pet, True)
    write(book, "PET lei-litru", "PET, doar lei/litru", PET_BRANDS, pet, False)
    shops_bib, have_bib = write(book, "Bag in box",
                                "Mărci la bag-in-box — preț de raft, SGR inclus",
                                BIB_BRANDS, bib, True)
    write(book, "BIB lei-litru", "Bag-in-box, doar lei/litru", BIB_BRANDS, bib, False)
    book.save(OUT)
    print(f"wrote {OUT}")
    print(f"  PET: {have_pet}/{len(PET_BRANDS)} mărci găsite, {shops_pet} magazine")
    print(f"  BIB: {have_bib}/{len(BIB_BRANDS)} mărci găsite, {shops_bib} magazine")


if __name__ == "__main__":
    main()
